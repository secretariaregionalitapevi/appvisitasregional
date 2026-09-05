import json
import logging
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from functools import wraps

import requests
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_http_methods

from .access_control import common_catalog
from .module_access import VALID_MODULES, invalidate_module_access
from .views import log_audit

logger = logging.getLogger(__name__)


ROLE_NAMES = {1: "Master", 2: "Admin", 3: "Coordenador", 4: "Instrutor"}
class ProfileUpdateConfigurationError(requests.RequestException):
    """Indica que a ponte RPC segura ainda n�o foi instalada no banco."""

ALLOWED_PROFILE_FIELDS = {
    "full_name", "username", "status", "role_id", "sector", "cargo", "comum",
    "municipio", "cidade", "cadastro_origem", "cadastro_origem_label",
    "cadastro_origem_rota", "cadastro_origem_setor_sugerido",
}


def _is_global(request):
    profile = request.session.get("user_profile") or {}
    try:
        if int(profile.get("role_id") or 99) == 1:
            return True
    except (TypeError, ValueError):
        pass
    return str(profile.get("role") or "").strip().upper() in {"MASTER", "GLOBAL"}


def global_only(view):
    @wraps(view)
    def wrapped(request, *args, **kwargs):
        if not _is_global(request):
            if request.path.startswith("/administracao/api/"):
                return JsonResponse({"error": "Acesso exclusivo para administradores globais."}, status=403)
            return redirect("/dashboard/v3")
        return view(request, *args, **kwargs)
    return wrapped


def _is_administrator(request):
    profile = request.session.get("user_profile") or {}
    try:
        return int(profile.get("role_id") or 99) in {1, 2}
    except (TypeError, ValueError):
        return str(profile.get("role") or "").strip().upper() in {"MASTER", "GLOBAL", "ADMIN"}


def administrative_only(view):
    @wraps(view)
    def wrapped(request, *args, **kwargs):
        if not _is_administrator(request):
            if "/api/" in request.path:
                return JsonResponse({"error": "Acesso exclusivo para administradores."}, status=403)
            return redirect("/dashboard/v3")
        return view(request, *args, **kwargs)
    return wrapped


def _headers(prefer=None):
    headers = {
        "apikey": settings.SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    return headers


def _get_table(table, params):
    response = requests.get(
        f"{settings.SUPABASE_URL}/rest/v1/{table}", headers=_headers(), params=params, timeout=15
    )
    response.raise_for_status()
    return response.json()


def _update_profile_direct(user_id, changes):
    """Compatibilidade no backend usando apenas os privilegios do service_role."""
    response = requests.patch(
        f"{settings.SUPABASE_URL}/rest/v1/profiles",
        headers=_headers("return=representation"),
        params={"user_id": f"eq.{user_id}"},
        json=changes,
        timeout=15,
    )
    response.raise_for_status()
    rows = response.json()
    if not isinstance(rows, list) or len(rows) != 1 or str(rows[0].get("user_id")) != str(user_id):
        raise requests.RequestException("O banco nao confirmou a atualizacao do perfil solicitado.")
    return rows[0]


def _update_profile_via_rpc(user_id, changes):
    """Usa a RPC, com compatibilidade para a migracao ausente ou claims antigas."""
    if not user_id or not isinstance(changes, dict) or not changes:
        raise ValueError("Informe o usuario e as alteracoes.")
    if set(changes) - (ALLOWED_PROFILE_FIELDS | {"role"}):
        raise ValueError("Campos de perfil nao permitidos.")
    if "status" in changes and changes["status"] not in {"pending", "approved", "rejected"}:
        raise ValueError("Status invalido.")
    changes = {key: (value.strip() or None) if isinstance(value, str) else value
               for key, value in changes.items()}
    response = requests.post(
        f"{settings.SUPABASE_URL}/rest/v1/rpc/admin_update_user_profile",
        headers=_headers(),
        json={"p_user_id": str(user_id), "p_changes": changes},
        timeout=15,
    )
    if response.status_code >= 400:
        try:
            error = response.json()
        except ValueError:
            error = {}
        if not isinstance(error, dict):
            error = {}
        missing_rpc = response.status_code == 404 and error.get("code") == "PGRST202"
        legacy_claims = (response.status_code == 403 and error.get("code") == "42501"
                         and error.get("message") == "Service role required")
        if missing_rpc or legacy_claims:
            # Nao concede privilegios: o banco continua exigindo a chave de
            # servico do backend e aplicando suas permissoes e triggers.
            logger.warning("RPC administrativa indisponivel; usando acesso service_role do backend.")
            return _update_profile_direct(user_id, changes)
    response.raise_for_status()
    updated = response.json()
    if not isinstance(updated, dict) or str(updated.get("user_id")) != str(user_id):
        raise requests.RequestException("A rotina administrativa nao retornou o perfil solicitado.")
    return updated

@global_only
def audit_center(request):
    return render(request, "pages/auditoria.html")


@global_only
@require_http_methods(["GET"])
def administration_data(request):
    try:
        queries = {
            "profiles": ("profiles", {"select": "*", "order": "created_at.desc"}),
            "logs": ("audit_logs", {"select": "*", "order": "created_at.desc", "limit": "1000"}),
            "sessions": ("audit_access_sessions", {"select": "*", "order": "started_at.desc", "limit": "500"}),
            "levels": ("access_levels", {"select": "*", "order": "level_order.asc"}),
            "modules": ("user_module_access", {"select": "user_id,module,active,granted_by,granted_at,revoked_at"}),
        }
        # Consultas independentes: a latencia de rede nao precisa ser somada.
        with ThreadPoolExecutor(max_workers=5) as executor:
            pending = {key: executor.submit(_get_table, table, params)
                       for key, (table, params) in queries.items()}
            profile_rows = pending["profiles"].result()
            profiles = list({str(row.get("user_id")): row for row in profile_rows if row.get("user_id")}.values())
            logs = pending["logs"].result()
            sessions = pending["sessions"].result()
            levels = pending["levels"].result()
            try:
                module_access = pending["modules"].result()
            except requests.RequestException:
                module_access = []
        return JsonResponse({"profiles": profiles, "logs": logs, "sessions": sessions, "access_levels": levels, "module_access": module_access})
    except requests.RequestException as exc:
        logger.exception("Falha ao consultar dados administrativos: %s", exc)
        return JsonResponse({"error": "Falha ao consultar os dados administrativos."}, status=502)


@global_only
@require_http_methods(["GET"])
def pending_access_count(request):
    """Retorna somente a quantidade necessária pelo alerta global do cabeçalho."""
    try:
        rows = _get_table("profiles", {"select": "user_id", "status": "eq.pending"})
        unique_user_ids = {str(row.get("user_id")) for row in rows if row.get("user_id")}
        return JsonResponse({"pending_count": len(unique_user_ids)})
    except requests.RequestException as exc:
        logger.exception("Falha ao consultar acessos pendentes: %s", exc)
        return JsonResponse({"error": "Falha ao consultar acessos pendentes."}, status=502)


def _auth_user_email(user_id):
    response = requests.get(
        f"{settings.SUPABASE_URL}/auth/v1/admin/users/{user_id}",
        headers=_headers(), timeout=10,
    )
    response.raise_for_status()
    return response.json().get("email")


def _audit_author(event):
    if not event:
        return None
    details = event.get("details") or {}
    actor = details.get("actor") or {}
    actor_id = actor.get("user_id") or details.get("actor_user_id") or event.get("user_id")
    name = actor.get("name") or details.get("actor_name")
    if not name and actor_id:
        rows = _get_table("profiles", {"select": "full_name,username", "user_id": f"eq.{actor_id}", "limit": "1"})
        if rows:
            name = rows[0].get("full_name") or rows[0].get("username")
    return {"name": name, "user_id": actor_id, "at": event.get("created_at")}


@global_only
@require_http_methods(["GET"])
def administration_user_metadata(request, user_id):
    target = str(user_id)
    try:
        with ThreadPoolExecutor(max_workers=3) as executor:
            email_job = executor.submit(_auth_user_email, target)
            registration_job = executor.submit(_get_table, "audit_logs", {
                "select": "user_id,created_at,details", "action": "eq.REGISTER",
                "or": f"(details->>user_id.eq.{target},details->>target_user_id.eq.{target})", "order": "created_at.asc", "limit": "1",
            })
            update_job = executor.submit(_get_table, "audit_logs", {
                "select": "user_id,created_at,details",
                "action": "in.(UPDATE_USER_ACCESS,UPDATE_USER_MODULE_ACCESS,USER_APPROVAL_APPROVED,USER_APPROVAL_REJECTED)",
                "or": f"(details->>target_user_id.eq.{target},details->>reviewed_user_id.eq.{target})",
                "order": "created_at.desc", "limit": "1",
            })
            email = email_job.result()
            registrations, updates = registration_job.result(), update_job.result()
        if not registrations and email:
            # Eventos legados identificavam o alvo pelo email. Exige proximidade
            # do horario de criacao e um unico evento, evitando recriacoes ambiguas.
            profiles = _get_table("profiles", {"select": "created_at", "user_id": f"eq.{target}", "limit": "1"})
            if profiles and profiles[0].get("created_at"):
                created = datetime.fromisoformat(profiles[0]["created_at"].replace("Z", "+00:00"))
                legacy = _get_table("audit_logs", {
                    "select": "user_id,created_at,details", "action": "eq.REGISTER",
                    "details->>email": f"eq.{email}",
                    "and": f"(created_at.gte.{(created - timedelta(minutes=5)).isoformat()},created_at.lte.{(created + timedelta(minutes=5)).isoformat()})",
                    "order": "created_at.asc", "limit": "2",
                })
                if len(legacy) == 1:
                    registrations = legacy
        return JsonResponse({
            "email": email,
            "registration": _audit_author(registrations[0]) if registrations else None,
            "last_update": _audit_author(updates[0]) if updates else None,
        })
    except requests.RequestException:
        logger.warning("Falha ao consultar metadados administrativos do usuario %s", target)
        return JsonResponse({"error": "Nao foi possivel consultar email e autoria. Feche e abra o cadastro para tentar novamente."}, status=502)


@global_only
@require_http_methods(["PATCH", "DELETE"])
def administration_user(request, user_id):
    if request.method == "DELETE":
        current_user_id = str(
            request.session.get("user_id")
            or (request.session.get("user_profile") or {}).get("user_id")
            or ""
        )
        if current_user_id == str(user_id):
            return JsonResponse({"error": "Você não pode excluir a própria conta enquanto está conectado."}, status=400)
        try:
            before = _get_table("profiles", {"select": "*", "user_id": f"eq.{user_id}", "limit": "1"})
            if not before:
                return JsonResponse({"error": "Usuário não encontrado."}, status=404)
            _update_profile_via_rpc(user_id, {"status": "rejected"})
            auth_response = requests.delete(
                f"{settings.SUPABASE_URL}/auth/v1/admin/users/{user_id}",
                headers=_headers(), timeout=15,
            )
            if auth_response.status_code not in {200, 204}:
                logger.error("Auth user deletion failed with HTTP %s", auth_response.status_code)
                return JsonResponse({"error": "Não foi possível excluir a conta de autenticação."}, status=502)
            # Mantém compatibilidade com bancos onde o perfil não usa CASCADE.
            profile_response = requests.delete(
                f"{settings.SUPABASE_URL}/rest/v1/profiles",
                headers=_headers(), params={"user_id": f"eq.{user_id}"}, timeout=15,
            )
            if profile_response.status_code not in {200, 204}:
                logger.error("Residual profile deletion failed with HTTP %s", profile_response.status_code)
                return JsonResponse({"error": "A conta foi removida, mas o perfil residual não pôde ser apagado."}, status=502)
            log_audit(request, "DELETE_USER", "ADMIN", {
                "target_user_id": str(user_id), "deleted_profile": before[0],
            })
            return JsonResponse({"status": "deleted", "message": "Usuário excluído com sucesso."})
        except requests.RequestException as exc:
            logger.exception("Falha ao excluir usuário %s: %s", user_id, exc)
            return JsonResponse({"error": "Falha ao excluir o usuário."}, status=502)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    changes = {key: value for key, value in body.items() if key in ALLOWED_PROFILE_FIELDS}
    if not changes:
        return JsonResponse({"error": "Nenhuma alteração permitida foi enviada."}, status=400)
    if "role_id" in changes:
        try:
            changes["role_id"] = int(changes["role_id"])
        except (TypeError, ValueError):
            return JsonResponse({"error": "Nível de acesso inválido."}, status=400)
        if changes["role_id"] not in ROLE_NAMES:
            return JsonResponse({"error": "Nível de acesso fora do padrão deste projeto."}, status=400)
        changes["role"] = ROLE_NAMES[changes["role_id"]]
        if changes["role_id"] == 1:
            changes["sector"] = "Global"
    if "status" in changes and changes["status"] not in {"pending", "approved", "rejected"}:
        return JsonResponse({"error": "Status inválido."}, status=400)
    try:
        before = _get_table("profiles", {"select": "*", "user_id": f"eq.{user_id}", "limit": "1"})
        if not before:
            return JsonResponse({"error": "Usuário não encontrado."}, status=404)
        candidate = {**before[0], **changes}
        candidate_role = int(candidate.get("role_id") or 4)
        if candidate_role == 4:
            candidate_common = str(candidate.get("comum") or "").strip()
            common_row = next(
                (item for item in common_catalog() if str(item.get("comum") or "").strip() == candidate_common),
                None,
            )
            if not common_row:
                return JsonResponse({
                    "error": "Para aprovar um acesso local, selecione uma comum válida."
                }, status=400)
            municipality = str(common_row.get("cidade") or "").strip()
            changes.update({"comum": candidate_common, "municipio": municipality, "cidade": municipality})
        updated = _update_profile_via_rpc(user_id, changes)
        log_audit(request, "UPDATE_USER_ACCESS", "ADMIN", {
            "target_user_id": user_id, "before": before[0], "changes": changes,
        })
        return JsonResponse(updated)
    except ProfileUpdateConfigurationError as exc:
        logger.error("Ponte RPC de atualiza��o ausente para o usu�rio %s: %s", user_id, exc)
        return JsonResponse({"error": str(exc)}, status=503)
    except requests.RequestException as exc:
        logger.exception("Falha ao atualizar usu�rio %s: %s", user_id, exc)
        return JsonResponse({"error": "Falha de comunica��o ao atualizar o usu�rio. Tente novamente em instantes."}, status=502)
OPERATIONAL_CONFIG = {
    "ministerio": {
        "table": "ministerio_regional",
        "required": {"nome", "ministerio", "comum", "municipio"},
        "fields": {"nome", "data_apresentacao", "data_ordenacao", "ministerio", "comum", "municipio", "administracao", "rrm", "aprovador_rrm", "cadastro_completo", "status", "possui_foto", "sexo", "observacoes"},
        "order": "nome.asc",
        "module": "MINISTERIO_REGIONAL",
    },
    "santa-ceia": {
        "table": "santa_ceia_eventos",
        "required": {"data_evento", "municipio", "comum"},
        "fields": {"data_evento", "municipio", "comum", "atendimento", "hora", "palavra", "oracao_pao", "oracao_calice", "diaconos", "ano_anterior"},
        "order": "data_evento.desc",
        "module": "SANTA_CEIA",
    },
}


def _request_json(request):
    try:
        value = json.loads(request.body or "{}")
    except json.JSONDecodeError as exc:
        raise ValueError("JSON invalido.") from exc
    if not isinstance(value, dict):
        raise ValueError("O corpo da requisicao deve ser um objeto.")
    return value


def _audit_changes(before, after, fields):
    return {
        key: {"anterior": (before or {}).get(key), "novo": (after or {}).get(key)}
        for key in fields if (before or {}).get(key) != (after or {}).get(key)
    }


@administrative_only
def administration(request):
    return redirect("/administracao/congregacoes/")


@administrative_only
def operational_page(request, section):
    if section not in {"congregacoes", "ministerio", "santa-ceia"}:
        return redirect("/administracao/congregacoes/")
    labels = {"congregacoes": "Congrega\u00e7\u00f5es", "ministerio": "Minist\u00e9rio", "santa-ceia": "Santa Ceia"}
    return render(request, "pages/administracao_operacional.html", {"section": section, "section_title": labels[section]})


def _catalog_norm(value):
    import unicodedata
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join("".join(char for char in text if not unicodedata.combining(char)).upper().split())


def _common_code(value):
    import re
    match = re.search(r"\bBR-\d{2}-\d{4}\b", str(value or ""), re.IGNORECASE)
    return match.group(0).upper() if match else ""


def _common_name(value):
    parts = [part.strip() for part in str(value or "").split(" - ") if part.strip()]
    return parts[-1] if len(parts) > 1 else str(value or "").strip()


def _congregation_rows():
    ministry = _get_table("ministerio_regional", {"select": "*", "order": "nome.asc", "limit": "3000"})
    by_code, by_name = {}, {}
    for person in ministry:
        common = str(person.get("comum") or "").strip()
        code = _common_code(common)
        if code:
            by_code.setdefault(code, []).append(person)
        by_name.setdefault(_catalog_norm(common), []).append(person)
    rows = []
    for common in common_catalog():
        full_name = str(common.get("comum") or common.get("nome") or "").strip()
        code = _common_code(full_name)
        servants = by_code.get(code, []) if code else by_name.get(_catalog_norm(full_name), [])
        rows.append({
            "id": common.get("id") or code or full_name,
            "codigo": code,
            "nome": full_name,
            "nome_comum": _common_name(full_name),
            "municipio": common.get("cidade") or common.get("municipio") or next((x.get("municipio") for x in servants if x.get("municipio")), ""),
            "administracao": next((x.get("administracao") for x in servants if x.get("administracao")), ""),
            "quantidade_servos": len(servants), "servos": servants,
        })
    return sorted(rows, key=lambda row: (str(row.get("municipio") or ""), str(row.get("nome") or "")))


@administrative_only
@require_http_methods(["GET"])
def operational_data(request, section):
    try:
        if section == "congregacoes":
            return JsonResponse({"rows": _congregation_rows()})
        config = OPERATIONAL_CONFIG.get(section)
        if not config:
            return JsonResponse({"error": "Secao administrativa invalida."}, status=404)
        rows = _get_table(config["table"], {"select": "*", "order": config["order"], "limit": "2000"})
        payload = {"rows": rows}
        if section == "santa-ceia":
            payload["counts"] = _get_table("santa_ceia_contagem", {"select": "*", "order": "created_at.desc", "limit": "2000"})
        return JsonResponse(payload)
    except requests.RequestException as exc:
        logger.exception("Falha ao consultar %s: %s", section, exc)
        return JsonResponse({"error": "Nao foi possivel consultar os dados administrativos."}, status=502)


@administrative_only
@require_http_methods(["POST", "PATCH", "DELETE"])
def operational_record(request, section, record_id=None):
    config = OPERATIONAL_CONFIG.get(section)
    if not config:
        return JsonResponse({"error": "Secao administrativa invalida."}, status=404)
    try:
        before = []
        if record_id:
            before = _get_table(config["table"], {"select": "*", "id": f"eq.{record_id}", "limit": "1"})
            if not before:
                return JsonResponse({"error": "Registro nao encontrado."}, status=404)
        if request.method == "DELETE":
            response = requests.delete(f"{settings.SUPABASE_URL}/rest/v1/{config['table']}", headers=_headers(), params={"id": f"eq.{record_id}"}, timeout=15)
            response.raise_for_status()
            log_audit(request, "DELETE", config["module"], {"entity": config["table"], "entity_id": str(record_id), "before": before[0], "outcome": "success"})
            return JsonResponse({"status": "deleted"})
        body = _request_json(request)
        changes = {key: body.get(key) for key in config["fields"] if key in body}
        if not changes:
            return JsonResponse({"error": "Nenhum campo permitido foi enviado."}, status=400)
        candidate = {**(before[0] if before else {}), **changes}
        missing = [field for field in config.get("required", set()) if not str(candidate.get(field) or "").strip()]
        if missing:
            return JsonResponse({"error": "Preencha os campos obrigatorios: " + ", ".join(sorted(missing)) + "."}, status=400)
        for field in {"data_apresentacao", "data_ordenacao", "data_evento"} & changes.keys():
            value = changes[field]
            if value is None or (isinstance(value, str) and not value.strip()):
                changes[field] = None
            else:
                try:
                    changes[field] = date.fromisoformat(str(value).strip()).isoformat()
                except ValueError:
                    raise ValueError(f"Data invalida no campo {field}. Use AAAA-MM-DD.")
        if "ano_anterior" in changes:
            value = changes["ano_anterior"]
            try:
                changes["ano_anterior"] = None if value is None or str(value).strip() == "" else int(str(value).strip())
            except ValueError:
                raise ValueError("Informe um numero inteiro no campo ano_anterior.")
        for boolean_field in {"cadastro_completo", "possui_foto"} & changes.keys():
            changes[boolean_field] = str(changes[boolean_field]).lower() in {"1", "true", "on", "sim"}
        actor_id = request.session.get("user_id")
        if section == "ministerio":
            changes["updated_by"] = actor_id
            if request.method == "POST":
                changes["created_by"] = actor_id
        elif section == "santa-ceia":
            changes["user_id"] = actor_id
        url = f"{settings.SUPABASE_URL}/rest/v1/{config['table']}"
        if request.method == "POST":
            response = requests.post(url, headers=_headers("return=representation"), json=changes, timeout=15)
        else:
            response = requests.patch(url, headers=_headers("return=representation"), params={"id": f"eq.{record_id}"}, json=changes, timeout=15)
        response.raise_for_status()
        result = response.json()[0]
        event = "CREATE" if request.method == "POST" else "UPDATE"
        log_audit(request, event, config["module"], {
            "entity": config["table"], "entity_id": str(result.get("id") or record_id),
            "before": before[0] if before else None, "after": result,
            "changes": _audit_changes(before[0] if before else {}, result, config["fields"]), "outcome": "success",
        })
        return JsonResponse(result, status=201 if request.method == "POST" else 200)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except requests.RequestException as exc:
        logger.exception("Falha ao alterar %s: %s", section, exc)
        log_audit(request, "MUTATION_FAILED", config["module"], {"entity": config["table"], "entity_id": str(record_id or ""), "outcome": "failure", "error_type": type(exc).__name__})
        return JsonResponse({"error": "Nao foi possivel concluir a operacao."}, status=502)

@administrative_only
@require_http_methods(["GET"])
def export_congregations_excel(request):
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = _congregation_rows()
    municipality = str(request.GET.get("municipio") or "").strip()
    search = _catalog_norm(request.GET.get("q"))
    if municipality:
        rows = [row for row in rows if str(row.get("municipio") or "") == municipality]
    if search:
        rows = [row for row in rows if search in _catalog_norm(" ".join(str(row.get(key) or "") for key in ("codigo", "nome", "municipio", "administracao")))]
    profile = request.session.get("user_profile") or {}
    actor = profile.get("full_name") or profile.get("username") or "Usuario do Sistema"
    now = datetime.now()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CONGREGACOES"
    headers = ["C\u00f3digo", "Congrega\u00e7\u00e3o", "Munic\u00edpio", "Administra\u00e7\u00e3o", "Servos vinculados", "Situacao"]
    last_column = get_column_letter(len(headers))
    navy, pale, stripe = "1E4B7A", "EAF2F8", "F4F6F8"
    for row_number, text, size, color, fill in [
        (1, "CONGREGA\u00c7\u00c3O CRIST\u00c3 NO BRASIL", 15, "FFFFFF", navy),
        (2, "Regional Itapevi - S\u00e3o Paulo", 10, "FFFFFF", navy),
        (3, "ADMINISTRA\u00c7\u00c3O - CONGREGA\u00c7\u00d5ES", 12, navy, pale),
    ]:
        sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=len(headers))
        cell = sheet.cell(row_number, 1, text)
        cell.font = Font(size=size, bold=row_number != 2, color=color)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells("A4:C4"); sheet["A4"] = f"Recorte: {municipality or 'Todos os munic\u00edpios'} - {len(rows)} comuns"
    sheet.merge_cells("D4:F4"); sheet["D4"] = f"Emissao: {now:%d/%m/%Y %H:%M} - Responsavel: {actor}"
    sheet["D4"].alignment = Alignment(horizontal="right")
    for cell in sheet[4]: cell.font = Font(size=9, bold=True, color="536A7D")
    thin = Side(style="thin", color="C8D1DA")
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(6, column, label)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = Border(bottom=thin)
    for index, row in enumerate(rows, 7):
        values = [row.get("codigo"), row.get("nome_comum"), row.get("municipio"), row.get("administracao"), row.get("quantidade_servos"), "Com minist\u00e9rio" if row.get("quantidade_servos") else "Sem minist\u00e9rio"]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(index, column, value if value is not None else "")
            cell.border = Border(bottom=thin); cell.alignment = Alignment(vertical="center")
            if index % 2 == 0: cell.fill = PatternFill("solid", fgColor=stripe)
        sheet.cell(index, 5).alignment = Alignment(horizontal="center")
    for index, width in enumerate([17, 44, 26, 30, 18, 20], 1): sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A7"; sheet.auto_filter.ref = f"A6:F{max(6, sheet.max_row)}"; sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"; sheet.page_setup.fitToWidth = 1; sheet.page_setup.fitToHeight = 0; sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:6"; sheet.oddFooter.center.text = "Administra\u00e7\u00e3o - Congrega\u00e7\u00f5es - Regional Itapevi"; sheet.oddFooter.right.text = "Pagina &P de &N"
    stream = io.BytesIO(); workbook.save(stream)
    log_audit(request, "EXPORT", "ADMINISTRACAO_CONGREGACOES", {"format": "xlsx", "row_count": len(rows), "municipality": municipality or None, "outcome": "success"})
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Congregacoes_{now:%d-%m-%Y_%H-%M}.xlsx"'
    return response


@administrative_only
@require_http_methods(["GET"])
def export_ministerio_excel(request):
    """Exportação institucional do recorte exibido no painel de Ministério."""
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = _get_table("ministerio_regional", {"select": "*", "order": "nome.asc", "limit": "3000"})
    municipality = str(request.GET.get("municipio") or "").strip()
    query = _catalog_norm(request.GET.get("q"))
    if municipality:
        rows = [row for row in rows if str(row.get("municipio") or "") == municipality]
    if query:
        rows = [row for row in rows if query in _catalog_norm(" ".join(str(row.get(key) or "") for key in ("nome", "ministerio", "comum", "municipio", "administracao", "rrm")))]
    profile = request.session.get("user_profile") or {}
    actor = profile.get("full_name") or profile.get("username") or "Usuário do Sistema"
    now = datetime.now()
    workbook = Workbook(); sheet = workbook.active; sheet.title = "MINISTERIO"
    headers = ["Nome", "Ministério", "Comum", "Município", "Administração", "RRM", "Status", "Apresentação"]
    navy, pale, stripe = "1E4B7A", "EAF2F8", "F4F6F8"
    for row_number, text, size, color, fill in [
        (1, "CONGREGAÇÃO CRISTÃ NO BRASIL", 15, "FFFFFF", navy),
        (2, "Regional Itapevi - São Paulo", 10, "FFFFFF", navy),
        (3, "ADMINISTRAÇÃO - MINISTÉRIO", 12, navy, pale),
    ]:
        sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=len(headers))
        cell = sheet.cell(row_number, 1, text); cell.font = Font(size=size, bold=row_number != 2, color=color); cell.fill = PatternFill("solid", fgColor=fill); cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells("A4:D4"); sheet["A4"] = f"Recorte: {municipality or 'Todos os municípios'} - {len(rows)} registros"
    sheet.merge_cells("E4:H4"); sheet["E4"] = f"Emissão: {now:%d/%m/%Y %H:%M} - Responsável: {actor}"; sheet["E4"].alignment = Alignment(horizontal="right")
    for cell in sheet[4]: cell.font = Font(size=9, bold=True, color="536A7D")
    thin = Side(style="thin", color="C8D1DA")
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(6, column, label); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=navy); cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = Border(bottom=thin)
    for index, row in enumerate(rows, 7):
        values = [row.get("nome"), row.get("ministerio"), row.get("comum"), row.get("municipio"), row.get("administracao"), row.get("rrm"), row.get("status"), row.get("data_apresentacao")]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(index, column, value if value is not None else ""); cell.border = Border(bottom=thin); cell.alignment = Alignment(vertical="center")
            if index % 2 == 0: cell.fill = PatternFill("solid", fgColor=stripe)
    for index, width in enumerate([33, 31, 40, 20, 22, 18, 14, 16], 1): sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A7"; sheet.auto_filter.ref = f"A6:H{max(6, sheet.max_row)}"; sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"; sheet.page_setup.fitToWidth = 1; sheet.page_setup.fitToHeight = 0; sheet.sheet_properties.pageSetUpPr.fitToPage = True; sheet.print_title_rows = "1:6"
    sheet.oddFooter.center.text = "Administração - Ministério - Regional Itapevi"; sheet.oddFooter.right.text = "Página &P de &N"
    stream = io.BytesIO(); workbook.save(stream)
    log_audit(request, "EXPORT", "MINISTERIO_REGIONAL", {"format": "xlsx", "row_count": len(rows), "municipality": municipality or None, "outcome": "success"})
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Ministerio_{now:%d-%m-%Y_%H-%M}.xlsx"'
    return response