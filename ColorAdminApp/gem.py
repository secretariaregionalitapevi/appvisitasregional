"""Painel gerencial do Grupo de Estudos Musicais (GEM)."""
import json
import io
import math
import re
import unicodedata
import uuid
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

import requests
from django.conf import settings
from django.core.cache import cache
from django.http import HttpResponse, JsonResponse
from django.http.multipartparser import MultiPartParser, MultiPartParserError
from django.shortcuts import render

from .access_control import can_access, filter_rows, scope_details, service_headers, user_scope
from .module_access import MODULE_MUSICALIZACAO, can_access_module
from .sam_history_sync import SOURCE_CONFIG, program_minimum_progress

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TABLE = "musica_acompanhamento_aluno"
SELECT_FIELDS = (
    "id,nome_aluno,status,comum_congregacao,cargo_ministerio,nivel,instrumento,"
    "municipio,programa_minimo_percentual,registro_msa,updated_at"
)
SUMMARY_FIELDS = "id,nome_aluno,status,nivel,instrumento,municipio,comum_congregacao,cargo_ministerio,registro_msa,programa_minimo_percentual,updated_at"
GRADUATION_TOKEN = "OFICIALIZAD"
LEVEL_OPTIONS = [
    'CANDIDATO(A)', 'CULTO OFICIAL', 'ENSAIO', 'MEIA HORA', 'OFICIALIZADO(A)', 'RJM',
    'RJM / CULTO OFICIAL', 'RJM / ENSAIO', 'RJM / MEIA HORA', 'RJM / OFICIALIZADO(A)',
]
INSTRUMENT_OPTIONS = ['\u00d3RG\u00c3O', 'ACORDEON', 'VIOLINO', 'VIOLA', 'VIOLONCELO', 'FLAUTA TRANSVERSAL', 'OBO\u00c9', "OBO\u00c9 D'AMORE", 'CORNE INGL\u00caS', 'CLARINETE', 'CLARINETE ALTO', 'CLARINETE BAIXO (CLARONE)', 'FAGOTE', 'SAXOFONE SOPRANO (RETO)', 'SAXOFONE ALTO', 'SAXOFONE TENOR', 'SAXOFONE BAR\u00cdTONO', 'TROMPETE', 'CORNET', 'FLUGELHORN', 'TROMPA', 'TROMBONE', 'TROMBONITO', 'BAR\u00cdTONO (PISTO)', 'EUF\u00d4NIO', 'TUBA']
MINISTRY_OPTIONS = ['M\u00daSICO', 'ORGANISTA']
TONALITY_OPTIONS = ['D\u00d3', 'F\u00c1', 'F\u00c1 / SI\u266d', 'L\u00c1', 'MI\u266d', 'SI\u266d']
STUDENT_CREATE_FIELDS = {'nome_aluno', 'comum_congregacao', 'municipio', 'cargo_ministerio', 'nivel', 'instrumento', 'possui_instrumento', 'instrumento_proprio', 'tonalidade', 'data_inicio_gem', 'data_nascimento', 'estado_civil', 'telefone', 'nome_responsavel', 'grau_parentesco', 'consentimento_lgpd'}


def _norm(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join("".join(char for char in text if not unicodedata.combining(char)).upper().split())


def is_graduated(row):
    """Oficialização encerra a formação, inclusive em níveis compostos."""
    return GRADUATION_TOKEN in _norm(row.get("nivel"))


def academic_status(row):
    return "graduado" if is_graduated(row) else "formacao"


def _percent(value):
    try:
        return max(0, min(100, int(round(float(value or 0)))))
    except (TypeError, ValueError):
        return 0


def _program_progress(student, msa_rows):
    return program_minimum_progress(msa_rows)


def _fetch_students():
    cached = cache.get("gem:students:v5")
    if cached is not None:
        return cached

    page_size = 1000
    url = f"{settings.SUPABASE_URL}/rest/v1/{TABLE}"
    count_response = requests.get(
        url, headers=service_headers("count=exact"),
        params={"select": "id", "limit": 1}, timeout=15,
    )
    count_response.raise_for_status()
    try:
        total = int(count_response.headers.get("Content-Range", "0/0").rsplit("/", 1)[-1])
    except ValueError:
        total = 10000

    def fetch_page(offset):
        response = requests.get(
            url,
            headers=service_headers(),
            params={"select": SUMMARY_FIELDS, "offset": offset, "limit": page_size},
            timeout=20,
        )
        response.raise_for_status()
        return offset, response.json()

    offsets = list(range(0, total, page_size))
    with ThreadPoolExecutor(max_workers=min(10, max(1, len(offsets)))) as executor:
        pages = sorted(executor.map(fetch_page, offsets), key=lambda item: item[0])
    rows = [row for _, page in pages for row in page]
    cache.set("gem:students:v5", rows, 300)
    return rows


def _visible_students(request):
    scope = user_scope(request)
    rows = []
    for source in _fetch_students():
        row = dict(source)
        row["comum"] = row.get("comum_congregacao")
        row["cidade"] = row.get("municipio")
        row["situacao_academica"] = academic_status(row)
        row["programa_minimo_informado"] = row.get("programa_minimo_percentual") is not None
        row["programa_minimo_percentual"] = _percent(row.get("programa_minimo_percentual"))
        rows.append(row)
    return scope, filter_rows(scope, rows)


def _event_date(value):
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return text[:10]


def operational_status_from_days(inactive_days):
    if inactive_days is None:
        return "SEM HISTORICO"
    if inactive_days > 180:
        return "INATIVO"
    if inactive_days > 90:
        return "ALERTA"
    return "ATIVO"


def _operational_activity(datasets, today=None):
    """Classifica presença pedagógica exclusivamente pelos lançamentos do SAM."""
    date_fields = {
        "msa": "data_aula", "metodo": "data_inicio", "hinario": "data",
        "provas": "data_prova", "escalas": "data", "atividades": "data_atividade",
    }
    activity_dates = []
    for source, field in date_fields.items():
        for row in datasets.get(source, []) or []:
            value = _event_date(row.get(field) or row.get("created_at"))
            try:
                activity_dates.append(datetime.fromisoformat(value).date())
            except (TypeError, ValueError):
                continue
    if not activity_dates:
        return {"last_activity_at": None, "inactive_days": None, "operational_status": "SEM HISTÓRICO", "requires_review": False}
    last_activity = max(activity_dates)
    inactive_days = ((today or datetime.now().date()) - last_activity).days
    status = operational_status_from_days(inactive_days)
    return {"last_activity_at": last_activity.isoformat(), "inactive_days": inactive_days,
            "operational_status": status, "requires_review": inactive_days > 365}


def _build_timeline(student, datasets):
    """Converte todas as fontes pedagógicas em uma cronologia única."""
    specs = {
        "msa": ("data_aula", "MSA", "fa-book-open", "fase", "paginas", "licoes"),
        "metodo": ("data_inicio", "Método", "fa-music", "metodo", "pagina", "licao"),
        "hinario": ("data", "Hinário", "fa-book-bible", "hino", "voz", None),
        "provas": ("data_prova", "Prova", "fa-clipboard-check", "modulo", "nota", None),
        "escalas": ("data", "Escala", "fa-wave-square", "escala", None, None),
    }
    events = []
    if student.get("created_at"):
        events.append({
            "date": _event_date(student["created_at"]), "type": "cadastro", "title": "Início do acompanhamento",
            "description": "Aluno incluído no acompanhamento do GEM.", "icon": "fa-user-plus", "source": "Cadastro",
        })
    for source, rows in datasets.items():
        if source == "atividades":
            for row in rows:
                events.append({
                    "date": _event_date(row.get("data_atividade") or row.get("created_at")),
                    "type": row.get("tipo_atividade") or "atividade",
                    "title": row.get("titulo") or "Marco acadêmico",
                    "description": row.get("descricao") or row.get("observacoes") or "",
                    "icon": "fa-flag-checkered", "source": "Marco",
                    "document_url": row.get("documento_url") or "",
                })
            continue
        date_field, label, icon, primary, secondary, tertiary = specs[source]
        for row in rows:
            details = []
            for field in (primary, secondary, tertiary):
                if field and row.get(field) not in (None, ""):
                    details.append(f"{field.replace('_', ' ').title()}: {row[field]}")
            if row.get("observacoes"):
                details.append(str(row["observacoes"]))
            events.append({
                "date": _event_date(row.get(date_field) or row.get("created_at")),
                "type": source, "title": label, "description": " · ".join(details),
                "icon": icon, "source": label, "authorized_by": row.get("autorizado_por") or "",
            })
    return sorted(events, key=lambda item: (item.get("date") or "", item.get("title") or ""), reverse=True)


def _milestones(level):
    normalized = _norm(level)
    current = 0
    if "OFICIALIZAD" in normalized:
        current = 4
    elif "CULTO OFICIAL" in normalized:
        current = 3
    elif "RJM" in normalized or "MEIA HORA" in normalized:
        current = 2
    elif "ENSAIO" in normalized:
        current = 1
    stages = [
        ("Início dos estudos", "Preparação e desenvolvimento inicial"),
        ("Ingresso no Ensaio", "Programa mínimo para ensaio"),
        ("Ingresso na RJM", "Apto para Reunião de Jovens e Menores"),
        ("Culto Oficial", "Apto para tocar nos cultos oficiais"),
        ("Oficialização", "Conclusão da formação musical"),
    ]
    return [{"title": title, "description": description, "status": "achieved" if index < current else "current" if index == current else "future"}
            for index, (title, description) in enumerate(stages)]


def can_open_module(request):
    # Nesta primeira etapa o GEM herda a autorização territorial da Musicalização.
    return can_access_module(request, MODULE_MUSICALIZACAO)


def _denied():
    return JsonResponse({"error": "Seu perfil não possui acesso à pasta GEM."}, status=403)


def page(request):
    if not can_open_module(request):
        return render(request, "pages/403.html", {"message": "Seu perfil não possui acesso à pasta GEM."}, status=403)
    profile = request.session.get('user_profile') or {}
    actor = profile.get('full_name') or profile.get('nome') or profile.get('name') or profile.get('email') or 'Usuario'
    return render(request, "pages/gem.html", {"scope": scope_details(user_scope(request)), 'report_user': actor})


def student_summary_page(request, student_id):
    if not can_open_module(request):
        return render(request, "pages/403.html", {"message": "Seu perfil não possui acesso à pasta GEM."}, status=403)
    return render(request, "pages/gem_student_summary.html", {"student_id": student_id})


def api_summary(request):
    if not can_open_module(request):
        return _denied()
    if request.method != "GET":
        return JsonResponse({"error": "Método não permitido."}, status=405)
    try:
        requested_scope = user_scope(request)
        summary_key = f"gem:summary:v5:{requested_scope['level']}:{_norm(requested_scope['municipio'])}:{_norm(requested_scope['comum'])}"
        cached_summary = cache.get(summary_key)
        if cached_summary is not None:
            return JsonResponse(cached_summary)
        scope, rows = _visible_students(request)
        formation = [row for row in rows if not is_graduated(row)]
        graduates = [row for row in rows if is_graduated(row)]
        complete = [row for row in formation if row["programa_minimo_percentual"] >= 100]
        with_progress = [row for row in formation if row["programa_minimo_informado"]]

        levels = Counter(_norm(row.get("nivel")) or "NÃO INFORMADO" for row in formation)
        instruments = Counter(_norm(row.get("instrumento")) or "A DEFINIR" for row in formation)
        municipalities = Counter(_norm(row.get("municipio")) or "NÃO INFORMADO" for row in formation)
        commons = Counter(_norm(row.get('comum_congregacao')) or 'NÃO INFORMADA' for row in formation)
        payload = {
            "scope": scope_details(scope),
            "totals": {
                "formation": len(formation),
                "graduates": len(graduates),
                "all": len(rows),
                "program_complete": len(complete),
                "program_tracked": len(with_progress),
                "average_progress": round(sum(row["programa_minimo_percentual"] for row in with_progress) / len(with_progress)) if with_progress else 0,
            },
            "levels": [{"label": label, "value": value} for label, value in levels.most_common()],
            "instruments": [{"label": label, "value": value} for label, value in instruments.most_common(8)],
            "instrument_options": sorted(instruments),
            "municipalities": [{"label": label, "value": value} for label, value in municipalities.most_common()],
            'commons': [{'label': label, 'value': value} for label, value in commons.most_common()],
            'catalogs': {'instruments': INSTRUMENT_OPTIONS, 'levels': LEVEL_OPTIONS, 'ministries': MINISTRY_OPTIONS, 'tonalities': TONALITY_OPTIONS},
        }
        cache.set(summary_key, payload, 300)
        return JsonResponse(payload)
    except (requests.RequestException, ValueError):
        return JsonResponse({"error": "Não foi possível consultar os dados do GEM."}, status=502)


def api_students(request):
    if not can_open_module(request):
        return _denied()
    if request.method == 'POST':
        try:
            raw = json.loads(request.body or '{}')
            payload = {key: raw.get(key) for key in STUDENT_CREATE_FIELDS if key in raw}
            required = ('nome_aluno', 'comum_congregacao', 'municipio', 'cargo_ministerio', 'nivel', 'instrumento')
            if any(not str(payload.get(key) or '').strip() for key in required):
                return JsonResponse({'error': 'Preencha todos os campos obrigatorios.'}, status=400)
            payload['nome_aluno'] = ' '.join(str(payload['nome_aluno']).upper().split())
            if len(payload['nome_aluno'].split()) < 2:
                return JsonResponse({'error': 'Informe o nome completo, sem abreviacoes.'}, status=400)
            if payload.get('instrumento') not in INSTRUMENT_OPTIONS:
                return JsonResponse({'error': 'Selecione um instrumento valido.'}, status=400)
            if payload.get('nivel') not in LEVEL_OPTIONS or payload.get('cargo_ministerio') not in MINISTRY_OPTIONS:
                return JsonResponse({'error': 'Nivel ou cargo invalido.'}, status=400)
            if payload.get('tonalidade') and payload['tonalidade'] not in TONALITY_OPTIONS:
                return JsonResponse({'error': 'Selecione uma tonalidade valida.'}, status=400)
            if payload.get('consentimento_lgpd') is not True:
                return JsonResponse({'error': 'O consentimento para tratamento dos dados e obrigatorio.'}, status=400)
            for boolean_field in ('possui_instrumento', 'instrumento_proprio'):
                if payload.get(boolean_field) not in (True, False, None):
                    return JsonResponse({'error': 'Opcao de instrumento invalida.'}, status=400)
            phone = re.sub(r'\D', '', str(payload.get('telefone') or ''))
            if phone and len(phone) not in (10, 11):
                return JsonResponse({'error': 'Informe um telefone valido com DDD.'}, status=400)
            if phone:
                payload['telefone'] = phone
            for date_field in ('data_nascimento', 'data_inicio_gem'):
                if payload.get(date_field):
                    try:
                        parsed_date = datetime.strptime(str(payload[date_field]), '%Y-%m-%d').date()
                    except ValueError:
                        return JsonResponse({'error': 'Informe as datas em formato valido.'}, status=400)
                    if parsed_date > datetime.now().date():
                        return JsonResponse({'error': 'As datas nao podem estar no futuro.'}, status=400)
            if payload.get('data_nascimento'):
                birth = datetime.strptime(str(payload['data_nascimento']), '%Y-%m-%d').date()
                today = datetime.now().date()
                age = today.year - birth.year - ((today.month, today.day) < (birth.month, birth.day))
                if age < 18 and not str(payload.get('nome_responsavel') or '').strip():
                    return JsonResponse({'error': 'Informe o responsavel para aluno menor de 18 anos.'}, status=400)
            candidate = dict(payload, comum=payload['comum_congregacao'], cidade=payload['municipio'])
            if not can_access(user_scope(request), candidate):
                return _denied()
            saved = requests.post(f"{settings.SUPABASE_URL}/rest/v1/{TABLE}", headers=service_headers('return=representation'), json=payload, timeout=20)
            saved.raise_for_status()
            cache.delete('gem:students:v5')
            return JsonResponse({'item': saved.json()[0]}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON invalido.'}, status=400)
        except (requests.RequestException, ValueError, IndexError):
            return JsonResponse({'error': 'Nao foi possivel cadastrar o aluno.'}, status=502)
    if request.method != "GET":
        return JsonResponse({"error": "Método não permitido."}, status=405)
    try:
        scope = user_scope(request)
        category = request.GET.get("situacao", "formacao")
        try:
            page_number = max(1, int(request.GET.get("page", 1)))
            page_size = max(10, min(100, int(request.GET.get("page_size", 25))))
        except ValueError:
            page_number, page_size = 1, 25
        params = {
            "select": SELECT_FIELDS, "order": "nome_aluno.asc",
            "offset": (page_number - 1) * page_size, "limit": page_size,
        }
        if category == "formacao":
            params["nivel"] = "not.ilike.*OFICIALIZAD*"
        elif category == "graduados":
            params["nivel"] = "ilike.*OFICIALIZAD*"
        if scope["level"] == "local":
            params["comum_congregacao"] = f"eq.{scope['comum']}"
        elif scope["level"] == "municipal":
            params["municipio"] = f"eq.{scope['municipio']}"
        for query_name, column in (("nivel", "nivel"), ("municipio", "municipio"), ('comum', 'comum_congregacao'), ("instrumento", "instrumento")):
            values = [value.strip() for value in request.GET.getlist(query_name) if value.strip()]
            if values:
                params[column] = f"eq.{values[0]}" if len(values) == 1 else 'in.(' + ','.join(json.dumps(value) for value in values) + ')'
        query = request.GET.get("q", "").strip()
        if query:
            safe_query = "".join(char for char in query if char not in "(),.*")[:80]
            if safe_query:
                params["or"] = "(" + ",".join(
                    f"{column}.ilike.*{safe_query}*" for column in
                    ("nome_aluno", "comum_congregacao", "municipio", "instrumento", "nivel")
                ) + ")"
        response = requests.get(
            f"{settings.SUPABASE_URL}/rest/v1/{TABLE}",
            headers=service_headers("count=exact"), params=params, timeout=15,
        )
        response.raise_for_status()
        rows = []
        for source in response.json():
            row = dict(source)
            row["comum"] = row.get("comum_congregacao")
            row["cidade"] = row.get("municipio")
            row["situacao_academica"] = academic_status(row)
            row["programa_minimo_informado"] = row.get("programa_minimo_percentual") is not None
            row["programa_minimo_percentual"] = _percent(row.get("programa_minimo_percentual"))
            rows.append(row)
        try:
            total = int(response.headers.get("Content-Range", "0/0").rsplit("/", 1)[-1])
        except ValueError:
            total = len(rows)
        return JsonResponse({
            "items": rows,
            "pagination": {
                "page": page_number,
                "page_size": page_size,
                "pages": max(1, math.ceil(total / page_size)),
                "total": total,
            },
        })
    except (requests.RequestException, ValueError):
        return JsonResponse({"error": "Não foi possível consultar os alunos do GEM."}, status=502)


def _report_rows(request):
    _scope, rows = _visible_students(request)
    category = request.GET.get('situacao', 'formacao')
    selected = {key: set(request.GET.getlist(key)) for key in ('nivel', 'municipio', 'comum', 'instrumento')}
    query = _norm(request.GET.get('q'))
    result = []
    for row in rows:
        if category == 'formacao' and is_graduated(row):
            continue
        if category == 'graduados' and not is_graduated(row):
            continue
        if selected['nivel'] and row.get('nivel') not in selected['nivel']:
            continue
        if selected['municipio'] and row.get('municipio') not in selected['municipio']:
            continue
        if selected['comum'] and row.get('comum_congregacao') not in selected['comum']:
            continue
        if selected['instrumento'] and row.get('instrumento') not in selected['instrumento']:
            continue
        searchable = ' '.join(str(row.get(field) or '') for field in ('nome_aluno', 'comum_congregacao', 'municipio', 'instrumento', 'nivel'))
        if query and query not in _norm(searchable):
            continue
        result.append(row)
    return sorted(result, key=lambda row: _norm(row.get('nome_aluno')))


def api_students_report(request):
    if not can_open_module(request):
        return _denied()
    if request.method != 'GET':
        return JsonResponse({'error': 'Metodo nao permitido.'}, status=405)
    return JsonResponse({'items': _report_rows(request)})


def export_students_excel(request):
    if not can_open_module(request):
        return _denied()
    rows = _report_rows(request)
    profile = request.session.get('user_profile') or {}
    actor = profile.get('full_name') or profile.get('nome') or profile.get('name') or profile.get('email') or 'Usuario'
    now = datetime.now()
    cities = request.GET.getlist('municipio')
    commons = request.GET.getlist('comum')
    scope = f"Municipios: {', '.join(cities) if cities else 'Todos'} | Comuns: {', '.join(commons) if commons else 'Todas'}"
    headers = ['Aluno', 'Registro MSA', 'Municipio', 'Comum congregacao', 'Cargo/Ministerio', 'Instrumento', 'Nivel', 'Programa minimo', 'Atualizacao']
    workbook = Workbook(); sheet = workbook.active; sheet.title = 'ALUNOS GEM'
    navy, pale = '1E4B7A', 'EAF2F8'; last = get_column_letter(len(headers)); thin = Side(style='thin', color='CCD5DD')
    for row_number, text_value, size in ((1, 'CONGREGAÇÃO CRISTÃ NO BRASIL', 15), (2, 'Regional Itapevi - São Paulo', 10), (3, 'GRUPO DE ESTUDOS MUSICAIS', 12)):
        sheet.merge_cells(f'A{row_number}:{last}{row_number}'); cell = sheet.cell(row_number, 1, text_value)
        cell.font = Font(size=size, bold=row_number != 2, color='FFFFFF' if row_number < 3 else navy)
        cell.fill = PatternFill('solid', fgColor=navy if row_number < 3 else pale); cell.alignment = Alignment(horizontal='center')
    sheet.merge_cells('A4:E4'); sheet['A4'] = 'Relatorio completo de alunos | ' + scope
    sheet.merge_cells(f'F4:{last}4'); sheet['F4'] = f'Emissao: {now:%d/%m/%Y %H:%M} | Responsavel: {actor}'; sheet['F4'].alignment = Alignment(horizontal='right')
    for cell in sheet[4]: cell.font = Font(size=8, bold=True, color='536A7D')
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(6, column, label); cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor=navy); cell.alignment = Alignment(horizontal='center'); cell.border = Border(bottom=thin)
    for row_number, row in enumerate(rows, 7):
        updated = str(row.get('updated_at') or '')
        try: updated = datetime.fromisoformat(updated.replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
        except ValueError: pass
        values = [row.get('nome_aluno'), row.get('registro_msa'), row.get('municipio'), row.get('comum_congregacao'), row.get('cargo_ministerio'), row.get('instrumento'), row.get('nivel'), f"{_percent(row.get('programa_minimo_percentual'))}%", updated]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value or ''); cell.border = Border(bottom=thin); cell.alignment = Alignment(vertical='center', wrap_text=column in (1, 4))
            if row_number % 2 == 0: cell.fill = PatternFill('solid', fgColor='F4F6F8')
    for index, width in enumerate([34, 17, 20, 38, 22, 25, 24, 16, 22], 1): sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = 'A7'; sheet.auto_filter.ref = f'A6:{last}{max(6, sheet.max_row)}'; sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = 'landscape'; sheet.page_setup.fitToWidth = 1; sheet.page_setup.fitToHeight = 0; sheet.sheet_properties.pageSetUpPr.fitToPage = True
    stream = io.BytesIO(); workbook.save(stream)
    response = HttpResponse(stream.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Relatorio_Alunos_GEM_{now:%d-%m-%Y_%H-%M}.xlsx'
    return response


def api_student_timeline(request, student_id):
    if not can_open_module(request):
        return _denied()
    if request.method != "GET":
        return JsonResponse({"error": "Método não permitido."}, status=405)
    try:
        scope = user_scope(request)
        student_response = requests.get(
            f"{settings.SUPABASE_URL}/rest/v1/{TABLE}", headers=service_headers(),
            params={"select": SELECT_FIELDS, "id": f"eq.{student_id}", "limit": 1}, timeout=15,
        )
        student_response.raise_for_status()
        source_rows = student_response.json()
        if not source_rows:
            # Não diferencia registro inexistente de registro fora do escopo.
            return JsonResponse({"error": "Aluno não encontrado neste escopo."}, status=404)
        student = dict(source_rows[0])
        student["comum"] = student.get("comum_congregacao")
        student["cidade"] = student.get("municipio")
        if not can_access(scope, student):
            return JsonResponse({"error": "Aluno não encontrado neste escopo."}, status=404)
        student["situacao_academica"] = academic_status(student)
        student["programa_minimo_informado"] = student.get("programa_minimo_percentual") is not None
        student["programa_minimo_percentual"] = _percent(student.get("programa_minimo_percentual"))

        sources = {
            "msa": ("musica_acompanhamento_msa", "data_aula.desc"),
            "metodo": ("musica_acompanhamento_metodo", "data_inicio.desc"),
            "hinario": ("musica_acompanhamento_hinario", "data.desc"),
            "provas": ("musica_acompanhamento_provas", "data_prova.desc"),
            "escalas": ("musica_acompanhamento_escala", "data.desc"),
            "atividades": ("musica_acompanhamento_atividades", "data_atividade.desc"),
        }

        def fetch(item):
            name, (table, order) = item
            response = requests.get(
                f"{settings.SUPABASE_URL}/rest/v1/{table}", headers=service_headers(),
                params={"select": "*", "aluno_id": f"eq.{student_id}", "order": order}, timeout=15,
            )
            response.raise_for_status()
            return name, response.json()

        with ThreadPoolExecutor(max_workers=6) as executor:
            datasets = dict(executor.map(fetch, sources.items()))
        student.update(_operational_activity(datasets))
        student["programa_minimo_percentual"] = _program_progress(student, datasets.get("msa"))
        student["programa_minimo_informado"] = bool(datasets.get("msa")) or student["programa_minimo_informado"]
        return JsonResponse({
            "student": student,
            "milestones": _milestones(student.get("nivel")),
            "events": _build_timeline(student, datasets),
            "counts": {name: len(rows) for name, rows in datasets.items()},
            # O resumo por abas precisa dos campos completos de cada registro,
            # não apenas da versão condensada usada pela antiga linha do tempo.
            "records": datasets,
        })
    except (requests.RequestException, ValueError):
        return JsonResponse({"error": "Não foi possível carregar a linha do tempo do aluno."}, status=502)


def api_student_detail(request, student_id):
    if not can_open_module(request):
        return _denied()
    if request.method != "PATCH":
        return JsonResponse({"error": "Método não permitido."}, status=405)
    fields = {"nome_aluno", "registro_msa", "comum_congregacao", "cargo_ministerio", "nivel", "instrumento", "municipio"}
    try:
        response = requests.get(
            f"{settings.SUPABASE_URL}/rest/v1/{TABLE}", headers=service_headers(),
            params={"select": "*", "id": f"eq.{student_id}", "limit": 1}, timeout=15,
        )
        response.raise_for_status()
        rows = response.json()
        if not rows:
            return JsonResponse({"error": "Aluno não encontrado."}, status=404)
        current = rows[0]
        current_location = dict(current, comum=current.get("comum_congregacao"), cidade=current.get("municipio"))
        if not can_access(user_scope(request), current_location):
            return _denied()
        raw = json.loads(request.body or "{}")
        payload = {key: raw.get(key) for key in fields if key in raw}
        candidate = dict(current, **payload, comum=payload.get("comum_congregacao", current.get("comum_congregacao")), cidade=payload.get("municipio", current.get("municipio")))
        if not payload or not can_access(user_scope(request), candidate):
            return _denied()
        saved = requests.patch(
            f"{settings.SUPABASE_URL}/rest/v1/{TABLE}", headers=service_headers("return=representation"),
            params={"id": f"eq.{student_id}"}, json=payload, timeout=15,
        )
        saved.raise_for_status()
        if payload.get("nivel") and payload["nivel"] != current.get("nivel"):
            normalized = _norm(payload["nivel"])
            milestone = ("oficializacao" if "OFICIALIZAD" in normalized else "ingresso_culto" if "CULTO OFICIAL" in normalized else "ingresso_rjm" if "RJM" in normalized or "MEIA HORA" in normalized else "ingresso_ensaio" if "ENSAIO" in normalized else None)
            if milestone:
                labels = {"ingresso_ensaio": "Ingresso no Ensaio", "ingresso_rjm": "Ingresso na RJM", "ingresso_culto": "Ingresso no Culto Oficial", "oficializacao": "Oficialização"}
                requests.post(f"{settings.SUPABASE_URL}/rest/v1/{SOURCE_CONFIG['atividades'][0]}", headers=service_headers("return=minimal"), json={"aluno_id": str(student_id), "tipo_atividade": milestone, "titulo": labels[milestone], "descricao": "Marco registrado automaticamente pela alteração de nível.", "data_atividade": datetime.now().date().isoformat(), "comum_congregacao": candidate.get("comum_congregacao"), "municipio": candidate.get("municipio")}, timeout=15).raise_for_status()
        cache.delete("gem:students:v5")
        return JsonResponse({"item": saved.json()[0]})
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except (requests.RequestException, ValueError, IndexError):
        return JsonResponse({"error": "Não foi possível atualizar o aluno."}, status=502)


def api_student_record(request, source_name, record_id=None):
    """Consulta e mantém um lançamento pedagógico, sempre validando o aluno e seu território."""
    if not can_open_module(request):
        return _denied()
    if source_name not in SOURCE_CONFIG:
        return JsonResponse({"error": "Tipo de lançamento inválido."}, status=404)
    if request.method not in {"GET", "POST", "PATCH", "DELETE"}:
        return JsonResponse({"error": "Método não permitido."}, status=405)
    table, editable_fields = SOURCE_CONFIG[source_name]
    try:
        multipart = request.content_type and request.content_type.startswith("multipart/form-data")
        uploaded_files = request.FILES
        if multipart and request.method == "PATCH":
            parser = MultiPartParser(request.META, request, request.upload_handlers, request.encoding)
            multipart_data, uploaded_files = parser.parse()
            raw = multipart_data.dict()
        elif multipart:
            raw = request.POST.dict()
        else:
            raw = json.loads(request.body or "{}") if request.method in {"POST", "PATCH"} else {}
        record = None
        if record_id:
            record_response = requests.get(
                f"{settings.SUPABASE_URL}/rest/v1/{table}", headers=service_headers(),
                params={"select": "*", "id": f"eq.{record_id}", "limit": 1}, timeout=15,
            )
            record_response.raise_for_status()
            rows = record_response.json()
            if not rows:
                return JsonResponse({"error": "Lançamento não encontrado."}, status=404)
            record = rows[0]
        student_id = (record or {}).get("aluno_id") or raw.get("aluno_id")
        if not student_id:
            return JsonResponse({"error": "Aluno não informado."}, status=400)
        student_response = requests.get(
            f"{settings.SUPABASE_URL}/rest/v1/{TABLE}", headers=service_headers(),
            params={"select": "id,comum_congregacao,municipio", "id": f"eq.{student_id}", "limit": 1}, timeout=15,
        )
        student_response.raise_for_status()
        students = student_response.json()
        if not students:
            return JsonResponse({"error": "Aluno vinculado não encontrado."}, status=404)
        student = dict(students[0], comum=students[0].get("comum_congregacao"), cidade=students[0].get("municipio"))
        if not can_access(user_scope(request), student):
            return _denied()
        if request.method == "GET":
            return JsonResponse({"item": record})
        if request.method in {"POST", "PATCH"}:
            payload = {key: raw.get(key) for key in editable_fields if key in raw}
            document = uploaded_files.get("documento") if multipart and source_name == "atividades" else None
            remove_document = source_name == "atividades" and str(raw.get("remover_documento", "")).lower() in {"1", "true", "yes"}
            if document:
                allowed = {"application/pdf", "image/jpeg", "image/png"}
                if document.content_type not in allowed or document.size > 10 * 1024 * 1024:
                    return JsonResponse({"error": "Documento inválido. Envie PDF, JPG ou PNG com até 10 MB."}, status=400)
                extension = {"application/pdf": ".pdf", "image/jpeg": ".jpg", "image/png": ".png"}[document.content_type]
                object_path = f"{student_id}/{uuid.uuid4().hex}{extension}"
                upload = requests.post(
                    f"{settings.SUPABASE_URL}/storage/v1/object/gem_documents/{object_path}",
                    headers={"apikey": settings.SUPABASE_SERVICE_ROLE_KEY, "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY}", "Content-Type": document.content_type},
                    data=document.read(), timeout=45,
                )
                upload.raise_for_status()
                payload["documento_url"] = object_path
                if not payload.get("nome_documento"):
                    payload["nome_documento"] = document.name
            elif remove_document:
                payload.update({"documento_url": None, "nome_documento": None})
            if not payload:
                return JsonResponse({"error": "Nenhum campo válido foi informado."}, status=400)
            if request.method == "POST":
                payload.update({"aluno_id": student_id, "comum_congregacao": student.get("comum_congregacao"), "municipio": student.get("municipio")})
                response = requests.post(
                    f"{settings.SUPABASE_URL}/rest/v1/{table}", headers=service_headers("return=representation"),
                    json=payload, timeout=15,
                )
            else:
                response = requests.patch(
                    f"{settings.SUPABASE_URL}/rest/v1/{table}", headers=service_headers("return=representation"),
                    params={"id": f"eq.{record_id}"}, json=payload, timeout=15,
                )
            response.raise_for_status()
            saved_item = response.json()[0]
            old_document_path = (record or {}).get("documento_url")
            if old_document_path and (remove_document or (document and old_document_path != saved_item.get("documento_url"))):
                try:
                    cleanup = requests.delete(
                        f"{settings.SUPABASE_URL}/storage/v1/object/gem_documents/{old_document_path}",
                        headers={"apikey": settings.SUPABASE_SERVICE_ROLE_KEY, "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY}"},
                        timeout=15,
                    )
                    cleanup.raise_for_status()
                except requests.RequestException:
                    pass
            return JsonResponse({"item": saved_item}, status=201 if request.method == "POST" else 200)
        response = requests.delete(
            f"{settings.SUPABASE_URL}/rest/v1/{table}", headers=service_headers("return=minimal"),
            params={"id": f"eq.{record_id}"}, timeout=15,
        )
        response.raise_for_status()
        return JsonResponse({}, status=204)
    except (json.JSONDecodeError, MultiPartParserError):
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except (requests.RequestException, ValueError, IndexError):
        return JsonResponse({"error": "Não foi possível atualizar o lançamento."}, status=502)


def activity_document(request, record_id):
    """Entrega documento privado somente após validar aluno e território."""
    if not can_open_module(request):
        return _denied()
    try:
        table = SOURCE_CONFIG["atividades"][0]
        response = requests.get(f"{settings.SUPABASE_URL}/rest/v1/{table}", headers=service_headers(), params={"select": "aluno_id,documento_url,nome_documento", "id": f"eq.{record_id}", "limit": 1}, timeout=15)
        response.raise_for_status(); rows = response.json()
        if not rows or not rows[0].get("documento_url"):
            return JsonResponse({"error": "Documento não encontrado."}, status=404)
        item = rows[0]
        student_response = requests.get(f"{settings.SUPABASE_URL}/rest/v1/{TABLE}", headers=service_headers(), params={"select": "id,comum_congregacao,municipio", "id": f"eq.{item['aluno_id']}", "limit": 1}, timeout=15)
        student_response.raise_for_status(); students = student_response.json()
        if not students or not can_access(user_scope(request), dict(students[0], comum=students[0].get("comum_congregacao"), cidade=students[0].get("municipio"))):
            return _denied()
        download = requests.get(f"{settings.SUPABASE_URL}/storage/v1/object/authenticated/gem_documents/{item['documento_url']}", headers={"apikey": settings.SUPABASE_SERVICE_ROLE_KEY, "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY}"}, timeout=30)
        download.raise_for_status()
        result = HttpResponse(download.content, content_type=download.headers.get("Content-Type", "application/octet-stream"))
        result["Content-Disposition"] = f'inline; filename="{item.get("nome_documento") or "documento"}"'
        return result
    except requests.RequestException:
        return JsonResponse({"error": "Não foi possível consultar o documento."}, status=502)
