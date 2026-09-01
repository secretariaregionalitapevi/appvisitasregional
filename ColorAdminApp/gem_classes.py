"""Painel gerencial de aulas e frequência alimentado pela sincronização incremental com o SAM."""
from collections import Counter, defaultdict
from datetime import datetime
import io
import json
import re
import unicodedata

import requests
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .access_control import can_access, common_catalog, scope_details, service_headers, user_scope
from .gem import can_open_module


def page(request):
    if not can_open_module(request):
        return render(request, "pages/403.html", {"message": "Seu perfil não possui acesso à pasta GEM."}, status=403)
    profile = request.session.get("user_profile") or {}
    actor = profile.get("full_name") or profile.get("nome") or profile.get("name") or profile.get("email") or "Usuário"
    return render(request, "pages/gem_classes.html", {"scope": scope_details(user_scope(request)), "report_user": actor})


def student_attendance_page(request, student_id):
    if not can_open_module(request):
        return render(request, "pages/403.html", {"message": "Seu perfil nao possui acesso a pasta GEM."}, status=403)
    profile = request.session.get("user_profile") or {}
    actor = profile.get("full_name") or profile.get("nome") or profile.get("name") or profile.get("email") or "Usuario"
    return render(request, "pages/gem_student_attendance.html", {"student_id": student_id, "report_user": actor})


def _location_key(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char)).upper()
    text = re.sub(r"^BR-\d+(?:-\d+)*\s*-\s*", "", text)
    return " ".join(text.split())


def _municipality_map():
    result = {}
    for row in common_catalog():
        common = row.get("comum") or ""
        city = row.get("cidade") or row.get("municipio") or ""
        if common and city:
            result[_location_key(common)] = city
    return result


def _common_catalog_map():
    """Indexa a descricao importada do SAM no nome oficial com codigo BR."""
    result = {}
    for row in common_catalog():
        common = str(row.get("comum") or "").strip()
        if common:
            result[_location_key(common)] = {"label": common, "municipio": row.get("cidade") or row.get("municipio") or ""}
    return result


def _common_options(rows):
    options = {row.get("congregacao"): row.get("congregacao_label") or row.get("congregacao") for row in rows if row.get("congregacao")}
    return [{"value": value, "label": label} for value, label in sorted(options.items(), key=lambda item: _location_key(item[1]))]

def _get(table, **params):
    response = requests.get(
        f"{settings.SUPABASE_URL}/rest/v1/{table}", headers=service_headers(), params=params, timeout=30,
    )
    response.raise_for_status()
    return response.json()


def _rate(rows):
    return round(sum(row.get("presente") is True for row in rows) * 100 / len(rows)) if rows else 0


def _grade(value):
    try:
        grade = float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None
    if 0 <= grade <= 10:
        return grade
    if 10 < grade <= 100:
        return grade / 10
    return None


def api_student_attendance(request, student_id):
    if not can_open_module(request):
        return JsonResponse({"error": "Seu perfil nao possui acesso a pasta GEM."}, status=403)
    if request.method != "GET":
        return JsonResponse({"error": "Metodo nao permitido."}, status=405)
    try:
        calls = _get(
            "sam_gem_attendance",
            select="aula_id,aluno_id,nome_aluno,presente,source_member_id",
            aluno_id=f"eq.{student_id}", limit=10000,
        )
        if not calls:
            return JsonResponse({"error": "Nenhuma chamada de frequencia foi encontrada para este aluno."}, status=404)

        students = _get(
            "musica_acompanhamento_aluno",
            select="id,nome_aluno,status,comum_congregacao,cargo_ministerio,nivel,instrumento,municipio,programa_minimo_percentual,registro_msa",
            id=f"eq.{student_id}", limit=1,
        )
        student_record = students[0] if students else {}

        lessons = []
        class_ids = list(dict.fromkeys(str(row["aula_id"]) for row in calls))
        for start in range(0, len(class_ids), 100):
            ids = class_ids[start:start + 100]
            lessons.extend(_get(
                "sam_gem_classes",
                select="id,source_id,data_aula,congregacao,curso,turma,instrutor_aula",
                id=f"in.({','.join(ids)})", limit=1000,
            ))

        scope = user_scope(request)
        city_by_common = _municipality_map()
        catalog_by_common = _common_catalog_map()
        visible_lessons = {}
        for lesson in lessons:
            common_key = _location_key(lesson.get("congregacao"))
            lesson["municipio"] = city_by_common.get(common_key, "Nao identificado")
            lesson["congregacao_label"] = (catalog_by_common.get(common_key) or {}).get("label") or lesson.get("congregacao")
            if can_access(scope, {"comum": lesson.get("congregacao"), "municipio": lesson.get("municipio")}):
                visible_lessons[str(lesson["id"])] = lesson
        calls = [row for row in calls if str(row.get("aula_id")) in visible_lessons]
        if not calls:
            return JsonResponse({"error": "Aluno nao encontrado neste escopo."}, status=404)

        exams = _get(
            "musica_acompanhamento_provas", select="*", aluno_id=f"eq.{student_id}",
            order="data_prova.desc", limit=1000,
        )
        exam_rows = []
        for exam in exams:
            grade = _grade(exam.get("nota"))
            exam_rows.append({
                "data_prova": exam.get("data_prova"), "modulo": exam.get("modulo") or "Prova",
                "nota": grade, "observacoes": exam.get("observacoes"),
                "situacao": "Aprovado" if grade is not None and grade >= 7 else "Reprovado" if grade is not None else "Sem nota",
            })
        valid_grades = [row["nota"] for row in exam_rows if row["nota"] is not None]
        exam_summary = {
            "quantidade": len(valid_grades),
            "media": round(sum(valid_grades) / len(valid_grades), 1) if valid_grades else None,
            "aprovadas": sum(grade >= 7 for grade in valid_grades),
            "reprovadas": sum(grade < 7 for grade in valid_grades),
            "nota_minima": 7,
        }

        rows = []
        monthly = defaultdict(list)
        for call in calls:
            lesson = visible_lessons[str(call["aula_id"])]
            row = {**lesson, "presente": call.get("presente") is True}
            rows.append(row)
            period = str(lesson.get("data_aula") or "")[:7]
            if period:
                monthly[period].append(call)
        rows.sort(key=lambda row: (row.get("data_aula") or "", row.get("curso") or ""), reverse=True)
        present = sum(row["presente"] for row in rows)
        latest = rows[0]
        chronological = sorted(rows, key=lambda row: row.get("data_aula") or "")
        first_date = datetime.fromisoformat(chronological[0]["data_aula"][:10]).date()
        for exam in exam_rows:
            try:
                exam_date = datetime.fromisoformat(str(exam.get("data_prova") or "")[:10]).date()
                exam_month_distance = (exam_date.year - first_date.year) * 12 + exam_date.month - first_date.month
                exam["semestre"] = min(4, max(1, exam_month_distance // 6 + 1))
            except ValueError:
                exam["semestre"] = None
        semester_rows = [[] for _ in range(4)]
        for row in chronological:
            lesson_date = datetime.fromisoformat(row["data_aula"][:10]).date()
            month_distance = (lesson_date.year - first_date.year) * 12 + lesson_date.month - first_date.month
            semester_rows[min(3, max(0, month_distance // 6))].append(row)
        semesters = []
        for index, items in enumerate(semester_rows, 1):
            semester_present = sum(item["presente"] for item in items)
            semester_rate = round(semester_present * 100 / len(items)) if items else None
            semester_grades = [exam["nota"] for exam in exam_rows if exam.get("semestre") == index and exam.get("nota") is not None]
            semesters.append({
                "semestre": index, "aulas": len(items), "presencas": semester_present,
                "ausencias": len(items) - semester_present, "aproveitamento": semester_rate,
                "media_provas": round(sum(semester_grades) / len(semester_grades), 1) if semester_grades else None,
                "provas": len(semester_grades),
                "situacao": "Consolidado" if items and semester_rate >= 75 else "Em atencao" if items else "A cursar",
            })
        attendance_rate = round(present * 100 / len(rows))
        grade_average = exam_summary["media"]
        if attendance_rate >= 85 and (grade_average is None or grade_average >= 8):
            projection = ("ALTA", "Frequencia e notas indicam boa perspectiva de concluir os quatro semestres com aproveitamento.")
        elif attendance_rate >= 75 and (grade_average is None or grade_average >= 7):
            projection = ("ADEQUADA", "O aluno atende aos minimos atuais de frequencia e provas; deve manter regularidade ate o quarto semestre.")
        else:
            projection = ("BAIXA", "Frequencia ou media de provas abaixo do minimo exige plano de acompanhamento da gestao.")
        return JsonResponse({
            "student": {
                "id": str(student_id), "nome": student_record.get("nome_aluno") or calls[0].get("nome_aluno"),
                "comum": student_record.get("comum_congregacao") or latest.get("congregacao_label"),
                "municipio": student_record.get("municipio") or latest.get("municipio"),
                "instrumento": student_record.get("instrumento"), "nivel": student_record.get("nivel"),
                "cargo_ministerio": student_record.get("cargo_ministerio"), "status": student_record.get("status"),
                "registro_msa": student_record.get("registro_msa"),
                "programa_minimo_percentual": student_record.get("programa_minimo_percentual"),
            },
            "totals": {"aulas": len(rows), "presencas": present, "ausencias": len(rows) - present, "frequencia": attendance_rate},
            "semesters": semesters,
            "projection": {"nivel": projection[0], "mensagem": projection[1], "meta_frequencia": 75},
            "exams": exam_rows,
            "exam_summary": exam_summary,
            "months": [{"periodo": key, "aulas": len(monthly[key]), "frequencia": _rate(monthly[key])} for key in sorted(monthly)],
            "lessons": rows,
        })
    except requests.RequestException:
        return JsonResponse({"error": "Nao foi possivel consultar o boletim de frequencia."}, status=502)


def export_student_attendance_excel(request, student_id):
    if not can_open_module(request):
        return JsonResponse({"error": "Seu perfil nao possui acesso a pasta GEM."}, status=403)
    dashboard = api_student_attendance(request, student_id)
    if dashboard.status_code != 200:
        return dashboard
    data = json.loads(dashboard.content)
    student, totals = data["student"], data["totals"]
    profile = request.session.get("user_profile") or {}
    actor = profile.get("full_name") or profile.get("nome") or profile.get("name") or profile.get("email") or "Usuario"
    now = datetime.now()
    workbook = Workbook()
    navy, pale, alternate, thin_color = "1E4B7A", "EAF2F8", "F4F6F8", "CCD5DD"
    thin = Side(style="thin", color=thin_color)

    def prepare(sheet, title, headers):
        last = get_column_letter(len(headers))
        for row, text, fill, color, size in (
            (1, "CONGREGAÇÃO CRISTÃ NO BRASIL", navy, "FFFFFF", 15),
            (2, "Regional Itapevi - São Paulo | GRUPO DE ESTUDOS MUSICAIS", navy, "FFFFFF", 10),
            (3, title, pale, navy, 12),
        ):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            cell = sheet.cell(row, 1, text)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(size=size, bold=row != 2, color=color)
            cell.alignment = Alignment(horizontal="center")
        split = max(1, len(headers) // 2)
        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=split)
        sheet.cell(4, 1, f"Aluno: {student.get('nome') or '-'} | {student.get('comum') or '-'}")
        sheet.merge_cells(start_row=4, start_column=split + 1, end_row=4, end_column=len(headers))
        sheet.cell(4, split + 1, f"Emissão: {now:%d/%m/%Y %H:%M} | Impresso por: {actor}")
        sheet.cell(4, split + 1).alignment = Alignment(horizontal="right")
        for cell in sheet[4]:
            cell.font = Font(size=9, bold=True, color="536A7D")
        for column, label in enumerate(headers, 1):
            cell = sheet.cell(6, column, label)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        sheet.freeze_panes = "A7"
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = "1:6"
        return last

    summary = workbook.active
    summary.title = "BOLETIM"
    summary_headers = ["Semestre", "Aulas", "Presenças", "Ausências", "Frequência", "Provas", "Média das provas", "Situação"]
    summary_last = prepare(summary, "BOLETIM INDIVIDUAL DE DESEMPENHO", summary_headers)
    for row_index, semester in enumerate(data.get("semesters") or [], 7):
        values = [f"{semester['semestre']}o semestre", semester["aulas"], semester["presencas"], semester["ausencias"],
                  "A cursar" if semester["aproveitamento"] is None else f"{semester['aproveitamento']}%", semester.get("provas") or 0,
                  "Sem nota" if semester.get("media_provas") is None else f"{semester['media_provas']:.1f}", semester["situacao"]]
        for column, value in enumerate(values, 1):
            cell = summary.cell(row_index, column, value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=alternate)
    info_row = 12
    summary.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=8)
    summary.cell(info_row, 1, "DADOS DO ALUNO").fill = PatternFill("solid", fgColor=pale)
    summary.cell(info_row, 1).font = Font(bold=True, color=navy)
    student_lines = [
        ("Nome", student.get("nome")), ("Instrumento", student.get("instrumento")),
        ("Nivel", student.get("nivel")), ("Cargo/Ministerio", student.get("cargo_ministerio")),
        ("Comum", student.get("comum")), ("Municipio", student.get("municipio")),
        ("Programa minimo", f"{student.get('programa_minimo_percentual') or 0}%"),
        ("Frequencia acumulada", f"{totals['frequencia']}%"),
        ("Projecao de conclusao", data["projection"]["nivel"]), ("Leitura gerencial", data["projection"]["mensagem"]),
    ]
    for offset, (label, value) in enumerate(student_lines, info_row + 1):
        summary.cell(offset, 1, label).font = Font(bold=True, color=navy)
        summary.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=8)
        summary.cell(offset, 2, value or "Nao informado")
        summary.cell(offset, 2).alignment = Alignment(wrap_text=True)
    summary.auto_filter.ref = f"A6:{summary_last}{10}"
    for index, width in enumerate([18, 12, 14, 14, 17, 12, 20, 24], 1):
        summary.column_dimensions[get_column_letter(index)].width = width

    calls_sheet = workbook.create_sheet("CHAMADAS")
    call_headers = ["Data", "Semestre", "Comum / Turma", "Modalidade", "Instrutor", "Situação"]
    call_last = prepare(calls_sheet, "HISTORICO INDIVIDUAL DE CHAMADAS", call_headers)
    first_date = datetime.fromisoformat(data["lessons"][-1]["data_aula"][:10]).date()
    for row_index, row in enumerate(data.get("lessons") or [], 7):
        lesson_date = datetime.fromisoformat(row["data_aula"][:10]).date()
        semester = min(4, max(1, ((lesson_date.year - first_date.year) * 12 + lesson_date.month - first_date.month) // 6 + 1))
        values = [lesson_date.strftime("%d/%m/%Y"), f"{semester}o", f"{row.get('congregacao_label') or row.get('congregacao') or '-'} | {row.get('turma') or '-'}",
                  row.get("curso") or "Nao informado", row.get("instrutor_aula") or "Nao informado", "Presente" if row.get("presente") else "Ausente"]
        for column, value in enumerate(values, 1):
            cell = calls_sheet.cell(row_index, column, value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=column in (3, 5))
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=alternate)
    calls_sheet.auto_filter.ref = f"A6:{call_last}{max(6, calls_sheet.max_row)}"
    for index, width in enumerate([14, 12, 43, 24, 34, 16], 1):
        calls_sheet.column_dimensions[get_column_letter(index)].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    stamp = now.strftime("%d_%m_%Y")
    filename = f"Boletim_GEM_{_file_part(student.get('nome'))}_{stamp}.xlsx"
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def api_dashboard(request):
    if not can_open_module(request):
        return JsonResponse({"error": "Seu perfil não possui acesso à pasta GEM."}, status=403)
    if request.method != "GET":
        return JsonResponse({"error": "Método não permitido."}, status=405)
    try:
        classes = _get(
            "sam_gem_classes",
            select="id,source_id,data_aula,congregacao,curso,turma,instrutor_aula,synced_at",
            order="data_aula.desc", limit=1000,
        )
        scope = user_scope(request)
        city_by_common = _municipality_map()
        catalog_by_common = _common_catalog_map()
        for row in classes:
            catalog_common = catalog_by_common.get(_location_key(row.get("congregacao"))) or {}
            row["congregacao_label"] = catalog_common.get("label") or row.get("congregacao")
            row["municipio"] = city_by_common.get(_location_key(row.get("congregacao")), "Não identificado")
        classes = [row for row in classes if can_access(scope, {"comum": row.get("congregacao"), "municipio": row.get("municipio")})]
        available_municipalities = sorted({row.get("municipio") for row in classes if row.get("municipio") and row.get("municipio") != "Não identificado"})
        available_commons = _common_options(classes)
        available_courses = sorted({row.get("curso") for row in classes if row.get("curso")})
        selected_municipality = request.GET.get("municipio", "").strip()
        selected_common = request.GET.get("comum", "").strip()
        selected_course = request.GET.get("curso", "").strip()
        search = request.GET.get("q", "").strip().upper()
        if selected_municipality:
            classes = [row for row in classes if row.get("municipio") == selected_municipality]
        available_commons = _common_options(classes)
        if selected_common:
            classes = [row for row in classes if row.get("congregacao") == selected_common]
        if selected_course:
            classes = [row for row in classes if row.get("curso") == selected_course]
        attendance = []
        class_ids = [row["id"] for row in classes]
        for start in range(0, len(class_ids), 100):
            ids = class_ids[start:start + 100]
            attendance.extend(_get(
                "sam_gem_attendance",
                select="aula_id,aluno_id,nome_aluno,presente,source_member_id",
                aula_id=f"in.({','.join(ids)})", limit=10000,
            ))

        if search:
            matching_ids = {
                row["id"] for row in classes
                if search in " ".join(str(row.get(key) or "") for key in ("congregacao", "curso", "turma", "instrutor_aula")).upper()
            }
            matching_ids.update(
                row["aula_id"] for row in attendance if search in str(row.get("nome_aluno") or "").upper()
            )
            classes = [row for row in classes if row["id"] in matching_ids]
            attendance = [row for row in attendance if row["aula_id"] in matching_ids]
        by_class, by_student = defaultdict(list), defaultdict(list)
        class_map = {row["id"]: row for row in classes}
        for row in attendance:
            by_class[row["aula_id"]].append(row)
            by_student[str(row.get("aluno_id") or row.get("source_member_id"))].append(row)

        recent, classes_without_call, low_classes = [], 0, 0
        for lesson in classes:
            calls = by_class.get(lesson["id"], [])
            present = sum(item.get("presente") is True for item in calls)
            rate = round(present * 100 / len(calls)) if calls else None
            classes_without_call += not calls
            low_classes += rate is not None and rate < 75
            recent.append({
                **lesson, "total": len(calls), "presentes": present,
                "ausentes": len(calls) - present, "frequencia": rate,
            })

        risks = []
        for student_rows in by_student.values():
            if len(student_rows) < 2:
                continue
            present = sum(item.get("presente") is True for item in student_rows)
            rate = round(present * 100 / len(student_rows))
            if rate < 75:
                sample = student_rows[0]
                last_class = max(
                    (class_map.get(row["aula_id"], {}).get("data_aula") or "" for row in student_rows),
                    default="",
                )
                risks.append({
                    "aluno_id": sample.get("aluno_id"), "nome": sample.get("nome_aluno"),
                    "presencas": present, "ausencias": len(student_rows) - present,
                    "aulas": len(student_rows), "frequencia": rate, "ultima_aula": last_class,
                    "prioridade": "crítica" if rate < 50 else "atenção",
                })
        risks.sort(key=lambda row: (row["frequencia"], -row["aulas"], row["nome"] or ""))

        months = defaultdict(list)
        courses = defaultdict(list)
        for call in attendance:
            lesson = class_map.get(call["aula_id"], {})
            date = lesson.get("data_aula") or ""
            if len(date) >= 7:
                months[date[:7]].append(call)
            courses[lesson.get("curso") or "Não informado"].append(call)
        trend = [
            {"periodo": key, "frequencia": _rate(months[key]), "chamadas": len(months[key])}
            for key in sorted(months)[-8:]
        ]
        course_summary = sorted(
            ({"curso": key, "frequencia": _rate(rows), "chamadas": len(rows)}
             for key, rows in courses.items()),
            key=lambda row: (-row["chamadas"], row["curso"]),
        )[:8]
        commons = sorted(Counter(row.get("congregacao") or "Não informada" for row in classes).items())
        cursor = _get("sam_gem_sync_cursor", select="*", source="eq.aulas_abertas", limit=1)
        present_total = sum(row.get("presente") is True for row in attendance)
        linked = sum(bool(row.get("aluno_id")) for row in attendance)
        latest_date = max((row.get("data_aula") or "" for row in classes), default=None)
        return JsonResponse({
            "ready": True, "scope": scope_details(scope),
            "totals": {
                "aulas": len(classes), "chamadas": len(attendance), "presencas": present_total,
                "ausencias": len(attendance) - present_total, "frequencia": _rate(attendance),
                "alunos": len(by_student), "alunos_em_risco": len(risks),
                "aulas_abaixo_meta": low_classes, "aulas_sem_chamada": classes_without_call,
                "vinculados": linked, "conciliacao": round(linked * 100 / len(attendance)) if attendance else 0,
            },
            "recent": recent if request.GET.get("export") == "1" else recent[:120], "risk_students": risks if request.GET.get("export") == "1" else risks[:80], "trend": trend,
            "courses": course_summary,
            "commons": [{"label": label, "value": value} for label, value in commons],
            "filter_options": {"municipalities": available_municipalities, "commons": available_commons, "courses": available_courses},
            "filters": {"municipio": selected_municipality, "comum": selected_common, "curso": selected_course, "q": request.GET.get("q", "").strip()},
            "partial": not bool(cursor and cursor[0].get("last_full_sync_at")),
            "latest_class_at": latest_date,
            "sync": cursor[0] if cursor else None,
            "generated_at": datetime.now().isoformat(),
        })
    except requests.HTTPError as exc:
        if exc.response is not None and exc.response.status_code in (404, 400):
            return JsonResponse({"ready": False, "message": "Aplique a migração 018 para iniciar a sincronização de aulas do SAM."})
        return JsonResponse({"error": "Não foi possível consultar aulas e frequências."}, status=502)
    except requests.RequestException:
        return JsonResponse({"error": "Não foi possível consultar aulas e frequências."}, status=502)

def _report_date(value):
    if not value:
        return ""
    try:
        return datetime.fromisoformat(str(value)[:10]).strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


def _file_part(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(r"[^A-Za-z0-9]+", "_", "".join(c for c in text if not unicodedata.combining(c))).strip("_") or "Todos"


def export_excel(request):
    if not can_open_module(request):
        return JsonResponse({"error": "Seu perfil não possui acesso à pasta GEM."}, status=403)
    params = request.GET.copy()
    params["export"] = "1"
    request.GET = params
    dashboard = api_dashboard(request)
    if dashboard.status_code != 200:
        return dashboard
    data = json.loads(dashboard.content)
    profile = request.session.get("user_profile") or {}
    actor = profile.get("full_name") or profile.get("nome") or profile.get("name") or profile.get("email") or "Usuário"
    now = datetime.now()
    filters = data.get("filters") or {}
    scope = " · ".join(filter(None, [
        filters.get("municipio") or "Todos os municípios",
        filters.get("comum") or "Todas as comuns",
        filters.get("curso") or "Todas as modalidades",
    ]))
    workbook = Workbook()
    navy, pale, alternate = "1E4B7A", "EAF2F8", "F4F6F8"
    thin = Side(style="thin", color="CCD5DD")

    def prepare(sheet, title, headers):
        last = get_column_letter(len(headers))
        sheet.merge_cells(f"A1:{last}1")
        sheet["A1"] = "CONGREGAÇÃO CRISTÃ NO BRASIL"
        sheet["A1"].font = Font(size=15, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor=navy)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.merge_cells(f"A2:{last}2")
        sheet["A2"] = "Regional Itapevi - São Paulo"
        sheet["A2"].font = Font(size=10, color="FFFFFF")
        sheet["A2"].fill = PatternFill("solid", fgColor=navy)
        sheet["A2"].alignment = Alignment(horizontal="center")
        sheet.merge_cells(f"A3:{last}3")
        sheet["A3"] = "GRUPO DE ESTUDOS MUSICAIS · AULAS E FREQUÊNCIA"
        sheet["A3"].font = Font(size=12, bold=True, color=navy)
        sheet["A3"].fill = PatternFill("solid", fgColor=pale)
        sheet["A3"].alignment = Alignment(horizontal="center")
        split = max(1, len(headers) // 2)
        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=split)
        sheet.cell(4, 1, f"{title} · {scope}")
        if split < len(headers):
            sheet.merge_cells(start_row=4, start_column=split + 1, end_row=4, end_column=len(headers))
            sheet.cell(4, split + 1, f"Emissão: {now:%d/%m/%Y %H:%M} · Responsável: {actor}")
            sheet.cell(4, split + 1).alignment = Alignment(horizontal="right")
        for cell in sheet[4]:
            cell.font = Font(size=9, bold=True, color="536A7D")
        for column, label in enumerate(headers, 1):
            cell = sheet.cell(6, column, label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        sheet.freeze_panes = "A7"
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    lessons = workbook.active
    lessons.title = "AULAS"
    lesson_headers = ["Data", "Município", "Comum", "Modalidade / instrumento", "Turma", "Instrutor", "Presentes", "Ausentes", "Frequência"]
    prepare(lessons, "Histórico operacional de aulas", lesson_headers)
    for row_index, row in enumerate(data.get("recent") or [], 7):
        values = [_report_date(row.get("data_aula")), row.get("municipio"), row.get("congregacao"), row.get("curso"),
                  row.get("turma"), row.get("instrutor_aula"), row.get("presentes"), row.get("ausentes"),
                  f"{row.get('frequencia')}%" if row.get("frequencia") is not None else "Sem chamada"]
        for column, value in enumerate(values, 1):
            cell = lessons.cell(row_index, column, value if value is not None else "")
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=column in (3, 4, 5, 6))
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=alternate)
    for index, width in enumerate([14, 22, 38, 25, 35, 28, 12, 12, 14], 1):
        lessons.column_dimensions[get_column_letter(index)].width = width
    lessons.auto_filter.ref = f"A6:I{max(6, lessons.max_row)}"

    risks = workbook.create_sheet("ALUNOS EM ATENÇÃO")
    risk_headers = ["Aluno", "Aulas", "Presenças", "Ausências", "Frequência", "Prioridade", "Última aula"]
    prepare(risks, "Alunos abaixo da meta de 75%", risk_headers)
    for row_index, row in enumerate(data.get("risk_students") or [], 7):
        values = [row.get("nome"), row.get("aulas"), row.get("presencas"), row.get("ausencias"),
                  f"{row.get('frequencia')}%", str(row.get("prioridade") or "").title(), _report_date(row.get("ultima_aula"))]
        for column, value in enumerate(values, 1):
            cell = risks.cell(row_index, column, value if value is not None else "")
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=alternate)
    for index, width in enumerate([42, 12, 13, 13, 14, 16, 16], 1):
        risks.column_dimensions[get_column_letter(index)].width = width
    risks.auto_filter.ref = f"A6:G{max(6, risks.max_row)}"

    stream = io.BytesIO()
    workbook.save(stream)
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    file_scope = _file_part(filters.get("municipio") or "Todos_os_Municipios")
    response["Content-Disposition"] = f'attachment; filename="Relatorio_Frequencia_GEM_{file_scope}_{now:%d-%m-%Y_%H-%M}.xlsx"'
    return response
