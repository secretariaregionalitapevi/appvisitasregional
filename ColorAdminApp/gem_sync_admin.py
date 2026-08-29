"""Painel SAM. Classificações são somente informativas e nunca excluem alunos."""
import io
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone

import requests
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone as django_timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .access_control import can_access, scope_details, service_headers, user_scope
from .gem import can_open_module, operational_status_from_days


def _url(table):
    return f"{settings.SUPABASE_URL}/rest/v1/{table}"


def _file_part(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_") or "Todos_os_Municipios"


def _report_datetime(value):
    """Formato único dos relatórios; datas sem hora recebem 00:00."""
    if value in (None, ""):
        return ""
    text = str(value).strip()
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is not None:
            parsed = django_timezone.localtime(parsed)
        return parsed.strftime("%d/%m/%Y %H:%M")
    except ValueError:
        for pattern in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(text, pattern)
                return parsed.strftime("%d/%m/%Y %H:%M")
            except ValueError:
                continue
    return text


def _report_date(value):
    """Datas pedagógicas não exibem horário."""
    formatted = _report_datetime(value)
    return formatted[:10] if formatted else ""


def _get(table, params=None, limit=10000):
    response = requests.get(_url(table), headers=service_headers(), params={**(params or {}), "limit": limit}, timeout=(3.05, 30))
    response.raise_for_status()
    return response.json()


def _admin_allowed(request):
    return can_open_module(request) and user_scope(request)["level"] in {"global", "regional"}


def _runtime_control(control):
    heartbeat = control.get("heartbeat_at")
    heartbeat_age = None
    if heartbeat:
        try:
            heartbeat_age = max(0, int((datetime.now(timezone.utc) - datetime.fromisoformat(heartbeat.replace("Z", "+00:00"))).total_seconds()))
        except ValueError:
            pass
    worker_status = control.get("worker_status") or "unknown"
    online = heartbeat_age is not None and heartbeat_age < 45 and worker_status != "offline"
    desired = control.get("desired_state") or "paused"
    current_student = control.get("current_student")
    if not online:
        runtime_state = "queued" if desired == "running" else "offline"
    elif desired == "paused" and current_student:
        runtime_state = "stopping"
    elif desired == "paused":
        runtime_state = "paused"
    elif worker_status == "error":
        runtime_state = "error"
    elif worker_status == "starting":
        runtime_state = "starting"
    elif current_student or worker_status == "running":
        runtime_state = "running"
    else:
        runtime_state = "idle"
    return {**control, "worker_online": online, "heartbeat_age_seconds": heartbeat_age,
            "runtime_state": runtime_state, "can_start": runtime_state in {"offline", "paused", "error"},
            "can_pause": runtime_state in {"queued", "starting", "running", "idle"}}


def _denied():
    return JsonResponse({"error": "A administração do SAM exige perfil regional ou global."}, status=403)


@ensure_csrf_cookie
def page(request):
    if not _admin_allowed(request):
        return render(request, "pages/403.html", {"message": "A administração do SAM exige perfil regional ou global."}, status=403)
    profile = request.session.get("user_profile") or {}
    report_user = profile.get("full_name") or profile.get("nome") or profile.get("name") or profile.get("email") or "Usuário"
    return render(request, "pages/gem_sync_admin.html", {"scope": scope_details(user_scope(request)), "report_user": report_user})


def _visible_rows(request):
    rows = []
    offset = 0
    while True:
        batch = _get("sam_mirror_student_status", {"select": "*", "order": "source_name.asc", "offset": offset}, 1000)
        rows.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    missing_ids = {row.get("sync_state_id") for row in rows if not row.get("municipio")}
    if missing_ids:
        source_rows = []
        source_offset = 0
        while True:
            batch = _get("sam_student_sync_state", {
                "select": "id,source_common,source_instrument,source_level,source_payload",
                "aluno_id": "is.null", "offset": source_offset,
            }, 1000)
            source_rows.extend(batch)
            if len(batch) < 1000:
                break
            source_offset += 1000
        sources = {source.get("id"): source for source in source_rows if source.get("id") in missing_ids}
        for row in rows:
            source = sources.get(row.get("sync_state_id"))
            if not source:
                continue
            payload = source.get("source_payload") or {}
            row["municipio"] = payload.get("city") or row.get("municipio")
            row["comum_congregacao"] = source.get("source_common") or payload.get("common_name")
            row["instrumento"] = source.get("source_instrument") or payload.get("instrument")
            row["nivel"] = source.get("source_level") or payload.get("level")
            row["cargo_ministerio"] = payload.get("ministry") or row.get("cargo_ministerio")
    for row in rows:
        days = row.get("inactive_days")
        try:
            days = int(days) if days is not None else None
        except (TypeError, ValueError):
            days = None
        row["operational_status"] = operational_status_from_days(days)
        row["requires_review"] = days is not None and days > 365
    scope = user_scope(request)
    return [row for row in rows if can_access(scope, {"municipio": row.get("municipio"), "comum": row.get("comum_congregacao")})]


def api_dashboard(request):
    if not _admin_allowed(request):
        return _denied()
    try:
        rows = _visible_rows(request)
        controls = _get("sam_sync_control", {"select": "*", "id": "eq.1"}, 1)
        # O painel mostra somente cinco; o conjunto maior alimenta os relatórios.
        runs = _get("sam_sync_runs", {"select": "*", "order": "started_at.desc"}, 100)
        changes = _get("sam_level_history", {"select": "id,previous_level,new_level,effective_at,aluno_id", "order": "effective_at.desc"}, 12)
        statuses = Counter(row.get("operational_status") or "SEM HISTORICO" for row in rows)
        syncs = Counter(row.get("sync_status") or "pending" for row in rows)
        cities = defaultdict(lambda: Counter(total=0))
        for row in rows:
            city = row.get("municipio") or "NÃO INFORMADO"
            cities[city]["total"] += 1
            cities[city][row.get("operational_status") or "SEM HISTORICO"] += 1
            if row.get("requires_review"):
                cities[city]["EXCLUIR"] += 1
        control = controls[0] if controls else {}
        profile = request.session.get("user_profile") or {}
        profile_name = profile.get("full_name") or profile.get("nome") or profile.get("name")
        if profile_name and control.get("requested_by") == profile.get("email"):
            control["requested_by_display"] = profile_name
        runtime_control = _runtime_control(control)
        total = len(rows)
        synced = syncs["synced"]
        payload = {
            "totals": {"students": total, "synced": synced, "pending": syncs["pending"], "failed": syncs["failed"],
                       "unresolved": syncs["unmatched"] + syncs["ambiguous"], "progress": round(synced / total * 100, 1) if total else 0,
                       "active": statuses["ATIVO"], "alerts": statuses["ALERTA"] + statuses["INATIVO"],
                       "inactive": statuses["INATIVO"], "exclude": sum(1 for row in rows if row.get("requires_review")),
                       "no_history": statuses["SEM HISTORICO"]},
            "control": runtime_control,
            "statuses": dict(statuses), "sync_statuses": dict(syncs),
            "municipalities": [{"municipio": city, **counts} for city, counts in sorted(cities.items())],
            "runs": runs, "changes": changes,
        }
        return JsonResponse(payload)
    except requests.HTTPError as exc:
        status = getattr(exc.response, "status_code", 500)
        hint = " Aplique a migração 013_sam_mirror_admin.sql." if status in (400, 404) else ""
        return JsonResponse({"error": "Não foi possível carregar a administração do SAM." + hint}, status=503)
    except requests.RequestException:
        return JsonResponse({"error": "Supabase indisponível no momento."}, status=503)


@require_POST
def api_control(request):
    if not _admin_allowed(request):
        return _denied()
    try:
        import json
        action = json.loads(request.body or b"{}").get("action")
        if action not in {"start", "pause"}:
            return JsonResponse({"error": "Ação inválida."}, status=400)
        profile = (request.session.get("user_profile") or {})
        actor = profile.get("full_name") or profile.get("nome") or profile.get("name") or profile.get("email") or "usuário regional"
        now = datetime.now(timezone.utc).isoformat()
        response = requests.patch(_url("sam_sync_control"), headers=service_headers("return=representation"), params={"id": "eq.1"}, json={
            "desired_state": "running" if action == "start" else "paused", "requested_at": now,
            "requested_by": actor,
            "last_message": "Início solicitado; aguardando confirmação do worker" if action == "start" else "Pausa solicitada; concluindo o aluno atual com segurança",
            "updated_at": now,
        }, timeout=20)
        response.raise_for_status()
        return JsonResponse({"ok": True, "control": _runtime_control((response.json() or [{}])[0])})
    except requests.RequestException:
        return JsonResponse({"error": "Não foi possível enviar o comando ao worker."}, status=503)


def export_report(request):
    if not _admin_allowed(request):
        return _denied()
    rows = _visible_rows(request)
    city, status = request.GET.get("municipio", ""), request.GET.get("status", "")
    if city:
        rows = [r for r in rows if (r.get("municipio") or "") == city]
    if status:
        rows = [r for r in rows if (r.get("requires_review") if status == "EXCLUIR" else r.get("operational_status") == status)]
    profile = request.session.get("user_profile") or {}
    actor = profile.get("full_name") or profile.get("nome") or profile.get("name") or profile.get("email") or "Usuário"
    now = datetime.now()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MSA"
    headers = ["Nome", "Instrumento", "Localidade", "Cidade", "Cargo/Ministério", "Nível", "MSA Lançamento",
               "Fase MSA", "Status Geral", "Data da Verificação", "Observações"]
    last_column = get_column_letter(len(headers))
    navy, pale = "1E4B7A", "EAF2F8"
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = "CONGREGAÇÃO CRISTÃ NO BRASIL"
    sheet["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = "Regional Itapevi - São Paulo"
    sheet["A2"].font = Font(size=10, color="FFFFFF")
    sheet["A2"].fill = PatternFill("solid", fgColor=navy)
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(f"A3:{last_column}3")
    sheet["A3"] = "GRUPO DE ESTUDOS MUSICAIS · SAM"
    sheet["A3"].font = Font(size=12, bold=True, color=navy)
    sheet["A3"].fill = PatternFill("solid", fgColor=pale)
    sheet["A3"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(f"A4:F4")
    sheet["A4"] = f"Relatório de alunos · {city or 'Todos os municípios'}"
    sheet.merge_cells(f"G4:{last_column}4")
    sheet["G4"] = f"Emissão: {now:%d/%m/%Y %H:%M} · Impresso por: {actor}"
    sheet["G4"].alignment = Alignment(horizontal="right")
    for cell in sheet[4]:
        cell.font = Font(size=9, bold=True, color="536A7D")
    header_row = 6
    thin = Side(style="thin", color="CCD5DD")
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    status_colors = {"ATIVO": ("0B4568", "FFFFFF"), "ALERTA": ("FFF3DF", "B46B00"),
                     "INATIVO": ("A80000", "FFFFFF"), "EXCLUIR": ("FFFFFF", "D10000"),
                     "SEM HISTORICO": ("EDF2F5", "657B8C")}
    for row_index, row in enumerate(rows, header_row + 1):
        observations = row.get("last_msa_observations") or ""
        if row.get("requires_review"):
            observations = " · ".join(filter(None, ["REVISÃO / A EXCLUIR: mais de 365 dias sem atividade", observations]))
        values = [row.get("source_name"), row.get("instrumento"), row.get("comum_congregacao"), row.get("municipio"),
                  row.get("cargo_ministerio"), row.get("nivel"), _report_date(row.get("last_msa_date")), row.get("last_msa_phase"),
                  row.get("operational_status"), now.strftime("%d/%m/%Y %H:%M"), observations]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column, value or "")
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F4F6F8")
        status_cell = sheet.cell(row_index, 9)
        background, foreground = status_colors.get(row.get("operational_status"), ("FFFFFF", "263238"))
        status_cell.fill = PatternFill("solid", fgColor=background)
        status_cell.font = Font(bold=True, color=foreground)
        status_cell.alignment = Alignment(horizontal="center")
    widths = [38, 19, 48, 25, 20, 18, 16, 16, 18, 24, 32]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A6:{last_column}{max(header_row, sheet.max_row)}"
    sheet.row_dimensions[1].height = 24
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    summary = workbook.create_sheet("RELATÓRIO")
    summary.merge_cells("A1:D1")
    summary["A1"] = "CONGREGAÇÃO CRISTÃ NO BRASIL"
    summary["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=navy)
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.merge_cells("A2:D2")
    summary["A2"] = "Regional Itapevi - São Paulo · GRUPO DE ESTUDOS MUSICAIS · SAM"
    summary["A2"].font = Font(size=10, color="FFFFFF")
    summary["A2"].fill = PatternFill("solid", fgColor=navy)
    summary["A2"].alignment = Alignment(horizontal="center")
    summary.append([])
    summary.append(["Resumo de candidatos", "Quantidade", "%", "Observação"])
    counts = Counter((row.get("operational_status") or "SEM HISTORICO") for row in rows)
    review_count = sum(1 for row in rows if row.get("requires_review"))
    total = len(rows)
    labels = [("Total de candidatos", total, ""), ("Ativos", counts["ATIVO"], "ATIVO"),
              ("Alertas", counts["ALERTA"], "ALERTA"), ("Inativos", counts["INATIVO"], "INATIVO"),
              ("A excluir", review_count, "EXCLUIR"), ("Sem histórico", counts["SEM HISTORICO"], "SEM HISTORICO")]
    for label, amount, key in labels:
        summary.append([label, amount, (amount / total if total and key else None),
                        "Somente classificação informativa; não remove o aluno." if key == "EXCLUIR" else ""])
    summary.append([])
    summary.append(["Última atualização", now.strftime("%d/%m/%Y %H:%M"), "", f"Impresso por: {actor}"])
    for cell in summary[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center")
    for row_number in range(5, 11):
        summary.cell(row_number, 2).alignment = Alignment(horizontal="center")
        summary.cell(row_number, 3).number_format = "0.0%"
        for column in range(1, 5):
            summary.cell(row_number, column).border = Border(bottom=thin)
    for column, width in enumerate([28, 16, 13, 52], 1):
        summary.column_dimensions[get_column_letter(column)].width = width
    summary.sheet_view.showGridLines = False
    summary.page_setup.orientation = "landscape"
    summary.page_setup.fitToWidth = 1
    summary.page_setup.fitToHeight = 1
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    stream = io.BytesIO()
    workbook.save(stream)
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    file_scope = _file_part(city or "Todos os Municípios")
    response["Content-Disposition"] = f'attachment; filename="Relatorio_Alunos_SAM_{file_scope}_{now:%d-%m-%Y_%H-%M}.xlsx"'
    return response


def export_log_report(request):
    if not _admin_allowed(request):
        return _denied()
    runs = _get("sam_sync_runs", {"select": "*", "order": "started_at.desc"}, 1000)
    profile = request.session.get("user_profile") or {}
    actor = profile.get("full_name") or profile.get("nome") or profile.get("name") or profile.get("email") or "Usuário"
    now = datetime.now()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EXECUÇÕES"
    navy, pale = "1E4B7A", "EAF2F8"
    thin = Side(style="thin", color="CCD5DD")
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "CONGREGAÇÃO CRISTÃ NO BRASIL"
    sheet["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:H2")
    sheet["A2"] = "Regional Itapevi - São Paulo"
    sheet["A2"].font = Font(size=10, color="FFFFFF")
    sheet["A2"].fill = PatternFill("solid", fgColor=navy)
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A3:H3")
    sheet["A3"] = "GRUPO DE ESTUDOS MUSICAIS · SAM"
    sheet["A3"].font = Font(size=12, bold=True, color=navy)
    sheet["A3"].fill = PatternFill("solid", fgColor=pale)
    sheet["A3"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A4:D4")
    sheet["A4"] = "Relatório de execuções e logs"
    sheet.merge_cells("E4:H4")
    sheet["E4"] = f"Emissão: {now:%d/%m/%Y %H:%M} · Impresso por: {actor}"
    sheet["E4"].alignment = Alignment(horizontal="right")
    for cell in sheet[4]:
        cell.font = Font(size=9, bold=True, color="536A7D")
    headers = ["Início", "Término", "Estado", "Descobertos", "Alterados", "Eventos importados", "Erros", "Detalhes"]
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(6, column, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    status_names = {"success": "Concluído", "partial": "Parcial", "failed": "Falhou", "running": "Em execução"}
    for row_index, run in enumerate(runs, 7):
        details = run.get("details") or {}
        if not isinstance(details, str):
            import json
            details = json.dumps(details, ensure_ascii=False, sort_keys=True)
        values = [_report_datetime(run.get("started_at")), _report_datetime(run.get("finished_at")), status_names.get(run.get("status"), run.get("status")),
                  run.get("discovered_students"), run.get("changed_students"), run.get("imported_events"),
                  run.get("error_count"), details]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column, value if value is not None else "")
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=column == 8)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F4F6F8")
    for index, width in enumerate([24, 24, 16, 15, 14, 20, 12, 70], 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A6:H{max(6, sheet.max_row)}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    stream = io.BytesIO()
    workbook.save(stream)
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Relatorio_Logs_SAM_{now:%d-%m-%Y_%H-%M}.xlsx"'
    return response
