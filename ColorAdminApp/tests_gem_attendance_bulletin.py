import json
from io import BytesIO
from unittest.mock import patch
from uuid import UUID

from django.test import RequestFactory, SimpleTestCase
from django.http import JsonResponse
from django.urls import reverse
from openpyxl import load_workbook

from . import gem_classes


class GemAttendanceBulletinTests(SimpleTestCase):
    student_id = UUID("11111111-1111-1111-1111-111111111111")
    lesson_id = "22222222-2222-2222-2222-222222222222"

    def setUp(self):
        self.request = RequestFactory().get("/gem/aulas/api/alunos/boletim/")
        self.request.session = {"user_profile": {"full_name": "Gestor Teste"}}

    def test_bulletin_has_a_route_separate_from_msa_summary(self):
        url = reverse("ColorAdminApp:gemStudentAttendance", args=[self.student_id])
        self.assertEqual(url, f"/gem/aulas/alunos/{self.student_id}/boletim/")

    @patch("ColorAdminApp.gem_classes._common_catalog_map", return_value={"CENTRAL": {"label": "BR-22-0001 - CENTRAL"}})
    @patch("ColorAdminApp.gem_classes._municipality_map", return_value={"CENTRAL": "ITAPEVI"})
    @patch("ColorAdminApp.gem_classes.can_access", return_value=True)
    @patch("ColorAdminApp.gem_classes.user_scope", return_value={"level": "regional", "municipio": "", "comum": ""})
    @patch("ColorAdminApp.gem_classes.can_open_module", return_value=True)
    def test_api_builds_attendance_bulletin(self, _module, _scope, _access, _cities, _commons):
        def fake_get(table, **params):
            if table == "sam_gem_attendance":
                return [
                    {"aula_id": self.lesson_id, "aluno_id": str(self.student_id), "nome_aluno": "ANA", "presente": True, "source_member_id": 10},
                    {"aula_id": self.lesson_id, "aluno_id": str(self.student_id), "nome_aluno": "ANA", "presente": False, "source_member_id": 10},
                ]
            if table == "musica_acompanhamento_aluno":
                return [{"id": str(self.student_id), "nome_aluno": "ANA", "instrumento": "VIOLINO", "nivel": "CANDIDATA", "comum_congregacao": "BR-22-0001 - CENTRAL", "municipio": "ITAPEVI", "programa_minimo_percentual": 25}]
            if table == "musica_acompanhamento_provas":
                return [
                    {"data_prova": "2026-08-10", "modulo": "Teoria", "nota": "82,5", "observacoes": "Bom aproveitamento"},
                    {"data_prova": "2026-08-15", "modulo": "Pratica", "nota": 60, "observacoes": "Reforcar estudo"},
                ]
            return [{"id": self.lesson_id, "source_id": 20, "data_aula": "2026-08-20", "congregacao": "CENTRAL", "curso": "VIOLINO", "turma": "A", "instrutor_aula": "JOAO"}]

        with patch("ColorAdminApp.gem_classes._get", side_effect=fake_get):
            response = gem_classes.api_student_attendance(self.request, self.student_id)

        self.assertEqual(response.status_code, 200)
        body = json.loads(response.content)
        self.assertEqual(body["student"]["nome"], "ANA")
        self.assertEqual(body["student"]["comum"], "BR-22-0001 - CENTRAL")
        self.assertEqual(body["student"]["instrumento"], "VIOLINO")
        self.assertEqual(body["totals"], {"aulas": 2, "presencas": 1, "ausencias": 1, "frequencia": 50})
        self.assertEqual(len(body["semesters"]), 4)
        self.assertEqual(body["projection"]["nivel"], "BAIXA")
        self.assertEqual(body["exam_summary"], {"quantidade": 2, "media": 7.1, "aprovadas": 1, "reprovadas": 1, "nota_minima": 7})
        self.assertEqual(body["semesters"][0]["media_provas"], 7.1)
        self.assertEqual(body["exams"][0]["situacao"], "Aprovado")

    def test_grade_scale_accepts_seven_point_five(self):
        self.assertEqual(gem_classes._grade("7,5"), 7.5)
        self.assertEqual(gem_classes._grade("75"), 7.5)

    def test_students_panel_has_direct_bulletin_button(self):
        template = open("ColorAdminApp/templates/pages/gem.html", encoding="utf-8").read()
        self.assertIn('/gem/aulas/alunos/${esc(row.id)}/boletim/', template)
        self.assertIn("Boletim", template)

    @patch("ColorAdminApp.gem_classes.can_open_module", return_value=True)
    @patch("ColorAdminApp.gem_classes._get", return_value=[])
    def test_api_does_not_fall_back_to_unrelated_msa_history(self, _get, _module):
        response = gem_classes.api_student_attendance(self.request, self.student_id)
        self.assertEqual(response.status_code, 404)
        self.assertIn("chamada", json.loads(response.content)["error"])

    @patch("ColorAdminApp.gem_classes.can_open_module", return_value=True)
    def test_excel_uses_institutional_header_and_printable_sheets(self, _module):
        payload = {
            "student": {"nome": "ANA", "comum": "BR-22-0001 - CENTRAL", "municipio": "ITAPEVI", "instrumento": "VIOLINO", "nivel": "CANDIDATA", "cargo_ministerio": "MUSICA", "programa_minimo_percentual": 25},
            "totals": {"aulas": 2, "presencas": 1, "ausencias": 1, "frequencia": 50},
            "projection": {"nivel": "BAIXA", "mensagem": "Requer plano de recuperacao.", "meta_frequencia": 75},
            "semesters": [{"semestre": number, "aulas": 2 if number == 1 else 0, "presencas": 1 if number == 1 else 0, "ausencias": 1 if number == 1 else 0, "aproveitamento": 50 if number == 1 else None, "situacao": "Em atencao" if number == 1 else "A cursar"} for number in range(1, 5)],
            "lessons": [{"data_aula": "2026-08-20", "congregacao_label": "BR-22-0001 - CENTRAL", "turma": "A", "curso": "VIOLINO", "instrutor_aula": "JOAO", "presente": True}],
        }
        with patch("ColorAdminApp.gem_classes.api_student_attendance", return_value=JsonResponse(payload)):
            response = gem_classes.export_student_attendance_excel(self.request, self.student_id)
        self.assertRegex(response["Content-Disposition"], r"Boletim_GEM_ANA_\d{2}_\d{2}_\d{4}\.xlsx")
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["BOLETIM", "CHAMADAS"])
        for sheet in workbook.worksheets:
            self.assertEqual(sheet["A1"].value, "CONGREGAÇÃO CRISTÃ NO BRASIL")
            self.assertEqual(sheet.freeze_panes, "A7")
            self.assertEqual(sheet.page_setup.orientation, "landscape")
            self.assertEqual(sheet.page_setup.fitToWidth, 1)
            self.assertFalse(sheet.sheet_view.showGridLines)
            metadata = " | ".join(str(cell.value or "") for cell in sheet[4])
            self.assertIn("Impresso por: Gestor Teste", metadata)
        self.assertEqual(workbook["CHAMADAS"]["A7"].value, "20/08/2026")
