import io
from unittest.mock import patch

from django.test import RequestFactory, SimpleTestCase
from django.urls import resolve
from openpyxl import load_workbook

from .gem_sync_admin import _report_date, _report_datetime, export_log_report, export_report


class GemSyncAdminReportTests(SimpleTestCase):
    def test_report_dates_always_include_day_month_year_and_time(self):
        self.assertEqual(_report_datetime("2026-03-26"), "26/03/2026 00:00")
        self.assertEqual(_report_datetime("2026-08-29T11:48:00+00:00"), "29/08/2026 08:48")
        self.assertEqual(_report_date("2026-03-26"), "26/03/2026")

    @patch("ColorAdminApp.gem_sync_admin._visible_rows")
    @patch("ColorAdminApp.gem_sync_admin._admin_allowed", return_value=True)
    def test_student_excel_keeps_original_sheets_and_filename(self, _allowed, visible_rows):
        visible_rows.return_value = [{
            "source_name": "Aluno", "municipio": "ITAPEVI", "operational_status": "ATIVO",
            "last_msa_date": "2026-03-26",
        }]
        request = RequestFactory().get("/gem/sam/exportar/")
        request.session = {"user_profile": {"full_name": "Operador"}}

        response = export_report(request)
        workbook = load_workbook(io.BytesIO(response.content))

        self.assertEqual(workbook.sheetnames, ["MSA", "RELATÓRIO"])
        self.assertEqual(workbook["MSA"].max_column, 11)
        self.assertEqual([workbook["MSA"].cell(6, column).value for column in range(1, 12)], [
            "Nome", "Instrumento", "Localidade", "Cidade", "Cargo/Ministério", "Nível",
            "MSA Lançamento", "Fase MSA", "Status Geral", "Data da Verificação", "Observações",
        ])
        self.assertEqual(workbook["MSA"]["A6"].fill.fgColor.rgb, "001E4B7A")
        self.assertEqual(workbook["MSA"]["G7"].value, "26/03/2026")
        self.assertIn("Relatorio_Alunos_SAM", response["Content-Disposition"])

    @patch("ColorAdminApp.gem_sync_admin._get")
    @patch("ColorAdminApp.gem_sync_admin._admin_allowed", return_value=True)
    def test_log_excel_is_a_separate_download(self, _allowed, get_rows):
        get_rows.return_value = [{
            "started_at": "2026-08-29T11:00:00+00:00", "finished_at": "2026-08-29T11:10:00+00:00",
            "status": "partial", "discovered_students": 4451, "changed_students": 10,
            "imported_events": 42, "error_count": 3, "details": {"reason": "teste"},
        }]
        request = RequestFactory().get("/gem/sam/exportar-logs/")
        request.session = {"user_profile": {"full_name": "Operador"}}

        response = export_log_report(request)
        workbook = load_workbook(io.BytesIO(response.content))

        self.assertEqual(workbook.sheetnames, ["EXECUÇÕES"])
        self.assertEqual(workbook["EXECUÇÕES"]["C7"].value, "Parcial")
        self.assertEqual(workbook["EXECUÇÕES"]["A7"].value, "29/08/2026 08:00")
        self.assertEqual(workbook["EXECUÇÕES"]["G7"].value, 3)
        self.assertEqual(workbook["EXECUÇÕES"]["A6"].fill.fgColor.rgb, "001E4B7A")
        self.assertIn("Relatorio_Logs_SAM", response["Content-Disposition"])

    def test_both_excel_routes_are_registered_separately(self):
        self.assertEqual(resolve("/gem/sam/exportar/").url_name, "gemSyncExport")
        self.assertEqual(resolve("/gem/sam/exportar-logs/").url_name, "gemSyncLogExport")
